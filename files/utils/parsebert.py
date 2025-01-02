# General imports
import importlib
import sys
import os
import csv
import re
import warnings
import itertools
import pandas as pd
import torch

# External libraries for Excel reading
import xlrd
from xlrd.xldate import xldate_as_tuple, XLDateAmbiguous, XLDateError
from datetime import datetime, time
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter


#######################################################################################
## XLS Function


def process_xls( file_path, tokenizer, max_rows = 100, max_cols = 100, pad_length = 32 ):
    """
    Extracts metadata for each cell in the first sheet of the given spreadsheet and tokenizes the cell values 
    using a Hugging Face tokenizer (replacing the original vocab approach).

    Args:
        file_path (str): The path to the spreadsheet file.
        tokenizer: A Hugging Face tokenizer object (e.g., from transformers).
        max_rows (int): The maximum number of rows to process.
        max_cols (int): The maximum number of columns to process.
        pad_length (int): The length to which the tokens should be padded/truncated.

    Returns:
        x_tok (torch.Tensor): A 3D tensor of size (max_rows, max_cols, pad_length) containing tokenized cell values
            (i.e., input_ids).
        x_masks (torch.Tensor): A 3D tensor of the same shape as x_tok, containing attention masks.
        y_tok (torch.Tensor): A 3D tensor of size (max_rows, max_cols, 17) containing metadata for each cell:
            - Index 0: Cell type (int)
            - Index 1: Fill (int)
            - Index 2: Alignment (int)
            - Index 3: Font family (int)
            - Index 4: Style (int)
            - Index 5: Borders (int)
            - Index 6: Merge status (int)
    """

    def get_fill( xf_record, workbook ):
        """
        Determines whether the cell is filled with a background color.

        Args:
            xf_record (xlrd.formatting.XFRecord): The XFRecord object containing formatting information for the cell.
            workbook (xlrd.Book): The workbook object containing the cell.

        Returns:
            int: 1 if the cell is filled with a color other than white, 0 otherwise.
        """

        # Retrieve the color of the cell's background using the pattern_colour_index
        bg_color_rgb = workbook.colour_map.get(xf_record.background.pattern_colour_index)

        # Return 0 if the cell is not filled or filled with white; otherwise, return 1
        return 0 if bg_color_rgb in [ None, (255, 255, 255) ] else 1

    def get_merge( row, col, sheet ):
        """
        Determines the merge status of a cell in terms of its vertical and horizontal position within a merged range.

        Int to Merge Status Mapping:
            1 -> Outside any merged range ('O')
            2 -> Beginning of a merged range ('B')
            3 -> Inside a merged range but not at the beginning ('I')

        Args:
            row (int): The row index of the cell.
            col (int): The column index of the cell.
            sheet (xlrd.sheet.Sheet): The sheet object containing the merged cells information.

        Returns:
            int: A 2-digit integer where the first digit represents the vertical position
                 and the second digit represents the horizontal position within a merged range.
        """

        # Initialize tags as 0 (outside) by default
        vertical_tag = 0
        horizontal_tag = 0

        # Iterate over the merged cells in the sheet
        for start_row, end_row, start_col, end_col in sheet.merged_cells:

            # Check if the current cell is within a merged range
            if start_row <= row < end_row and start_col <= col < end_col:

                # Determine vertical tag
                if row == start_row:
                    vertical_tag = 1  # Beginning of vertical range
                else:
                    vertical_tag = 2  # Inside vertical range

                # Determine horizontal tag
                if col == start_col:
                    horizontal_tag = 1  # Beginning of horizontal range
                else:
                    horizontal_tag = 2  # Inside horizontal range

                # Break the loop as the cell is part of a merged range
                break

        # Return the combined vertical and horizontal tags
        return horizontal_tag, vertical_tag

    def get_data( cell, workbook, df, row, col ):
        """
        Determines the type of the cell, handles necessary conversions, and retrieves the cell value.

        Args:
            cell (xlrd.sheet.Cell): The cell object to determine the type and retrieve the value.
            workbook (xlrd.Book): The workbook object containing the cell.
            df (pandas.DataFrame): DataFrame representation of the sheet for value extraction.
            row (int): The row index of the cell.
            col (int): The column index of the cell.

        Returns:
            tuple: A tuple containing the determined type of the cell (int) and the cell value (str).
        """
        xlrd_error_map = {
            0: "#NULL!", 7: "#DIV/0!", 15: "#VALUE!", 23: "#REF!", 29: "#NAME?", 36: "#NUM!", 42: "#N/A", 43: "#GETTING_DATA",
        }

        # Empty cell with no formatting
        if cell.ctype == 0:
            return 0, ''

        # If text cell then check further to get exact text type
        elif cell.ctype == 1:
            format_str = workbook.format_map[ workbook.xf_list[ cell.xf_index ].format_key ].format_str

            if '#,##0' in format_str:
                return 6, str(cell.value)  # Currency type

            if format_str == '0.00E+00':
                return 7, f"{cell.value}"  # Scientific subclass of text

            return 1, str(cell.value)  # Text type

        # If numeric cell then check further to get exact numeric type
        elif cell.ctype == 2:
            format_str = workbook.format_map[ workbook.xf_list[ cell.xf_index ].format_key ].format_str

            if format_str == '0.00E+00':
                return 7, f"{cell.value:.2e}"  # Scientific subclass of number

            elif '#,##0' in format_str:

                if '$€' in format_str:
                    return 6, '€' + str(float(cell.value))

                return 6, '$' + str(float(cell.value))


            elif '%' in format_str:
                return 5, str(cell.value * 100) + '%'  # Percentage type

            elif isinstance(cell.value, float):
                if cell.value.is_integer():
                    return 3, str(df.iat[ row, col ])  # Integer type
                return 4, str(df.iat[ row, col ])  # Float type

            return 2, str(df.iat[ row, col ])  # General numeric type

        # If date cell then check further to get exact date type
        elif cell.ctype == 3:
            try:
                date_tuple = xldate_as_tuple(cell.value, workbook.datemode)

                if date_tuple[ :3 ] == (0, 0, 0):
                    return 9, str(df.iat[ row, col ])  # Time type

                elif date_tuple[ 3: ] == (0, 0, 0):
                    return 8, str(df.iat[ row, col ])  # Date type

                return 10, str(df.iat[ row, col ])  # Datetime type

            except (XLDateError, XLDateAmbiguous):
                return 10, str(df.iat[ row, col ])  # Datetime type

        # Boolean cell
        elif cell.ctype == 4:
            return 11, str(df.iat[ row, col ])

        # Error cell
        elif cell.ctype == 5:
            return 12, xlrd_error_map.get(cell.value, f"#ERR{cell.value}")

        # Blank cell (Empty cell with formatting)
        elif cell.ctype == 6:
            return 13, ''

        # Unknown type not in our keys return by default
        return 14, ''

    # Allocate x_tok, x_masks, y_tok first
    x_tok = torch.zeros((max_rows, max_cols, pad_length), dtype=torch.long)
    x_masks = torch.zeros((max_rows, max_cols, pad_length), dtype=torch.long)
    y_tok = torch.zeros((max_rows, max_cols, 17), dtype=torch.long)

    # Define trivial sequence for default x_tok and x_masks
    # We'll build a Python list for the minimal input_ids
    minimal_ids = [tokenizer.cls_token_id, tokenizer.sep_token_id] + [tokenizer.pad_token_id]*(pad_length - 2)
    minimal_mask = [1, 1] + [0]*(pad_length - 2)

    # Initialize 3D tensors of size row x col x seq = 100x100x<pad_length> for input_ids and masks
    # Use above sequences to set them by default
    for r in range(max_rows):
        for c in range(max_cols):
            x_tok[r, c, :] = torch.tensor(minimal_ids, dtype=torch.long)
            x_masks[r, c, :] = torch.tensor(minimal_mask, dtype=torch.long)

    # Open the .xls file with formatting_info=True to get the metadata
    workbook = xlrd.open_workbook(filename=file_path, formatting_info=True)

    # Apply the BIFF check to check for BIFF version
    if workbook.biff_version != 80:
        raise ValueError(f"{workbook.biff_version}: {file_path}")

    # Access the first sheet in the workbook
    sheet = workbook.sheet_by_index(0)

    # Get the df for some parsing
    df_read = pd.read_excel(file_path, header=None, dtype=str, na_values=' ', keep_default_na=False, engine='xlrd')

    # Convert to the size of df required
    df = df_read.reindex(index=range(max_rows), columns=range(max_cols), fill_value='')

    # Iterate over each combination of rows and columns in the sheet, up to the 100th row and column
    for row, col in itertools.product(range(min(sheet.nrows, max_rows)), range(min(sheet.ncols, max_cols))):

        # Retrieve the cell object from the sheet
        cell = sheet.cell(row, col)

        # Access the XFRecord object using the cell's XF index
        xf_record = workbook.xf_list[cell.xf_index]

        # Get cell train
        cell_type, cell_value = get_data(cell, workbook, df, row, col)

        # Encode the cell_value using huggingface tokenizer
        encoded_value = tokenizer.encode_plus(
            cell_value,
            max_length=pad_length,
            truncation=True,
            padding='max_length',
            return_tensors='pt'  # shape => [1, pad_length]
        )

        # Extract input_ids and attention_mask from encoded_value
        # and place them in x_tok and x_masks
        x_tok[row, col, :] = encoded_value['input_ids'][0]
        x_masks[row, col, :] = encoded_value['attention_mask'][0]

        # Set y_tok for current cell in y_tok
        y_tok[row, col, 0] = cell_type
        y_tok[row, col, 1] = get_fill(xf_record, workbook)  # Store 'fill'
        y_tok[row, col, 2] = xf_record.alignment.hor_align   # Store 'halign'
        y_tok[row, col, 3] = xf_record.alignment.vert_align  # Store 'valign'
        y_tok[row, col, 4] = workbook.font_list[xf_record.font_index].family   # Store 'font family'
        y_tok[row, col, 5] = workbook.font_list[xf_record.font_index].height // 20  # Store 'font size'
        y_tok[row, col, 6] = workbook.font_list[xf_record.font_index].bold    # Store 'bold'
        y_tok[row, col, 7] = workbook.font_list[xf_record.font_index].italic  # Store 'italic'
        y_tok[row, col, 8] = workbook.font_list[xf_record.font_index].underline_type  # Store 'underline'
        y_tok[row, col, 9] = workbook.font_list[xf_record.font_index].escapement      # Store 'escapement'
        rgb = workbook.colour_map.get(workbook.font_list[xf_record.font_index].colour_index)
        y_tok[row, col, 10] = 1 if (rgb and rgb != (0, 0, 0)) else 0  # Store 'colour'
        y_tok[row, col, 11] = xf_record.border.top_line_style       # Store top border style
        y_tok[row, col, 12] = xf_record.border.bottom_line_style
        y_tok[row, col, 13] = xf_record.border.left_line_style       # Store 'left border style'
        y_tok[row, col, 14] = xf_record.border.right_line_style      # Store 'right border style'
        y_tok[row, col, 15], y_tok[row, col, 16] = get_merge(row, col, sheet)  # Store 'hmerge' and 'vmerge'


        # If it is an empty/blank cell then further processing is required
        if y_tok[row, col, 0] in [0, 13, 14]:

            # Set font-related metadata to 0 (indices 4 to 10)
            y_tok[row, col, 4:11] = 0

            # If the cell is further of the type center across cells
            if y_tok[row, col, 2] == 6:

                # Define variable to store col of where this center continuous is starting
                start_col = col

                # Loop until we find a cell that is centered across cells but not empty or blank
                while y_tok[row, start_col, 2] == 6 and (y_tok[row, start_col, 0] == (0 or 13)):

                    # Move to the left
                    start_col = start_col - 1

                # Set the x_tok, x_masks, and y_tok tensor to these values
                #check_df.iloc[row, col] = check_df.iloc[row, start_col]
                x_tok[row, col, :] = x_tok[row, start_col, :].clone()
                x_masks[row, col, :] = x_masks[row, start_col, :].clone()
                y_tok[row, col, :] = y_tok[row, start_col, :].clone()

                # print(f'Cell: ({row}, {col})')  # ...
                # print(f"y_tok:\n{y_tok[row, col, :]}\n")

    # Iterate over the merged cells in the sheet
    for start_row, end_row, start_col, end_col in sheet.merged_cells:

        # # Print the merged cell locations for debugging
        # print(f'\n################NEW RANGE####################')
        # print(f'\nInitial Start Row: {start_row}, End Row: {end_row}, Start Col: {start_col}, End Col: {end_col},')

        # Derive adjusted indices for accessing the combined cell
        adj_sr = start_row
        adj_er = end_row - 1  # Cause end index is always one more than the actual merge
        adj_sc = start_col
        adj_ec = end_col - 1

        # Make sure to avoid out-of-bounds errors
        adj_sr = min(adj_sr, max_rows - 1)
        adj_sc = min(adj_sc, max_cols - 1)
        adj_er = min(adj_er, max_rows - 1)
        adj_ec = min(adj_ec, max_cols - 1)

        # Extract the starting values
        x_tok_start = x_tok[adj_sr, adj_sc, :].clone()
        x_masks_start = x_masks[adj_sr, adj_sc, :].clone()
        y_tok_start = y_tok[adj_sr, adj_sc, :].clone()
        #check_df_start = check_df.iloc[adj_sr, adj_sc]

        # Propagate values across the merged range
        for row, col in itertools.product(range(adj_sr, adj_er + 1), range(adj_sc, adj_ec + 1)):
            x_tok[row, col, :] = x_tok_start
            x_masks[row, col, :] = x_masks_start

            # Duplicate y_tok values excluding indices 15 and 16
            y_tok[row, col, :15] = y_tok_start[:15]
            y_tok[row, col, 17:] = y_tok_start[17:]

    # Return the 3D tensor of input_ids, attention masks, and the 3D metadata tensor
    return x_tok, x_masks, y_tok


#######################################################################################
## XLSX Function

def process_xlsx(
    file_path,
    tokenizer,                  # Replaces 'vocab'
    max_rows=100,
    max_cols=100,
    pad_length=32
):
    """
    Adapts the original process_xlsx to use a Hugging Face tokenizer 
    for tokenizing cell content, returning input_ids and attention masks
    in x_tok and x_masks, plus the metadata in y_tok.

    Args:
        file_path (str): Path to the .xlsx file.
        tokenizer: A Hugging Face tokenizer object (e.g., from transformers).
        max_rows (int): The maximum number of rows to process.
        max_cols (int): The maximum number of columns to process.
        pad_length (int): The length to which token sequences should be padded.

    Returns:
        x_tok (torch.Tensor): A 3D tensor of size (max_rows, max_cols, pad_length) 
                              containing tokenized cell values (input_ids).
        x_masks (torch.Tensor): A 3D tensor of the same shape as x_tok, 
                                containing attention masks.
        y_tok (torch.Tensor): A 3D tensor of size (max_rows, max_cols, 17) 
                              containing metadata for each cell.
    """

    def get_dataType(value, number_format):
        """
        Determines the type of the cell and handles necessary conversions.

        Int to Type Key Mapping:
            0  -> Empty cell
            1  -> Text cell
            2  -> Numeric type
            3  -> Integer subclass of number
            4  -> Float subclass of number
            5  -> Percentage subclass of number
            6  -> Currency subclass of number
            7  -> Scientific subclass of number
            8  -> Date type
            9  -> Time subclass of date
            10 -> Datetime subclass of date
            11 -> Boolean type with T or F value
            12 -> Error type corresponding to #REF!, #VALUE! etc. in excel
            13 -> Blank cell with formatting
            14 -> Unknown type not in our keys

        Args:
            value: The raw value of the cell in openpyxl.
            number_format (str): The cell's number format.

        Returns:
            int: The determined type of the cell (0-14).
        """

        # Empty cell with no value => blank
        if value is None:
            return 13

        # Boolean cell
        elif isinstance(value, bool):
            return 11

        # Error cell based on typical Excel error values
        elif any(error in str(value) for error in {
            "#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!"
        }):
            return 12

        # Currency format
        elif "#,##0" in number_format:
            return 6

        # Percentage format
        elif number_format.endswith("%"):
            return 5

        # Scientific notation format
        elif "E+" in number_format or "E-" in number_format:
            return 7

        # If float
        elif isinstance(value, float):
            if value.is_integer():
                return 3  # integer subclass
            return 4      # float

        # If int
        elif isinstance(value, int):
            return 3

        # If time type
        elif isinstance(value, time):
            return 9

        # If datetime type
        elif isinstance(value, datetime):
            # Check if time part is exactly midnight
            if value.time() == time(0, 0):
                # If the datetime is formatted as a date, return 8
                if "d" in number_format.lower() and "h" not in number_format.lower():
                    return 8
                else:
                    return 10
            else:
                return 10

        # Blank cell with formatting
        elif value == '':
            return 13

        # Text cell
        elif isinstance(value, str):
            return 1

        # Return 14 for unknown
        return 14

    def get_value(cell_type, cell):
        # If empty cell
        if cell_type == 0:
            return ''

        # Text cell
        elif cell_type == 1:
            return str(cell.value)

        # Numeric cell default
        elif cell_type == 2:
            return str(cell.value)

        # Integer subclass
        elif cell_type == 3:
            return str(cell.value)

        # Float subclass
        elif cell_type == 4:
            return str(cell.value)

        # Percentage
        elif cell_type == 5:
            return str(cell.value * 100) + '%'

        # Currency
        elif cell_type == 6:
            # If cell.value is already a string, might have a symbol
            if isinstance(cell.value, str):
                return str(cell.value)
            # Check if euro sign in format
            if '$€' in cell.number_format:
                return '€' + str(float(cell.value))
            # Otherwise, treat as USD
            return '$' + str(float(cell.value))

        # Scientific
        elif cell_type == 7:
            if isinstance(cell.value, (int, float)):
                return f"{cell.value:.2e}"
            else:
                return str(cell.value)

        # Date/time
        elif cell_type in [8, 9, 10]:
            return str(cell.value)

        # Boolean
        elif cell_type == 11:
            return str(cell.value)

        # Error
        elif cell_type == 12:
            return str(cell.value)

        # Blank
        elif cell_type == 13:
            return ''

        # Unknown
        return ''

    def get_alignment(cell):
        """
        Maps openpyxl alignment to custom numeric codes for horizontal/vertical alignment.
        """
        openpyxl_to_xlrd = {
            'horizontal': {
                'general': 0, 'left': 1, 'center': 2, 'right': 3,
                'fill': 4, 'justify': 5, 'centerContinuous': 6, 'distributed': 7
            },
            'vertical': {
                'top': 0, 'center': 1, 'bottom': 2, 'justify': 3, 'distributed': 4
            }
        }

        horiz = openpyxl_to_xlrd['horizontal'].get(cell.alignment.horizontal, 0)
        vert = openpyxl_to_xlrd['vertical'].get(cell.alignment.vertical, 2)
        return horiz, vert

    def get_borders(cell):
        """
        Maps openpyxl border styles to numeric codes.
        """
        openpyxl_to_xlrd = {
            'none': 0, 'thin': 1, 'medium': 2, 'dashed': 3, 'dotted': 4, 'thick': 5,
            'double': 6, 'hair': 7, 'mediumDashed': 8, 'dashDot': 9, 'mediumDashDot': 10,
            'dashDotDot': 11, 'mediumDashDotDot': 12, 'slantDashDot': 13
        }

        top = openpyxl_to_xlrd.get(cell.border.top.style or 'none')
        bot = openpyxl_to_xlrd.get(cell.border.bottom.style or 'none')
        left = openpyxl_to_xlrd.get(cell.border.left.style or 'none')
        right = openpyxl_to_xlrd.get(cell.border.right.style or 'none')
        return top, bot, left, right

    def get_merge(row, col, sheet):
        """
        Determines the merge status in openpyxl-based merges.

        Int to Merge Status Mapping:
            1 -> Outside any merged range ('O')
            2 -> Beginning of a merged range ('B')
            3 -> Inside a merged range but not at the beginning ('I')
        """
        vertical_tag = 0
        horizontal_tag = 0
        row_1_based = row + 1
        col_1_based = col + 1
        cell_coord = get_column_letter(col_1_based) + str(row_1_based)

        for merged in sheet.merged_cells.ranges:
            if cell_coord in merged:
                if row_1_based == merged.min_row:
                    vertical_tag = 1
                else:
                    vertical_tag = 2
                if col_1_based == merged.min_col:
                    horizontal_tag = 1
                else:
                    horizontal_tag = 2
                break

        return horizontal_tag, vertical_tag

    def get_fontcol(cell):
        """
        Determines whether the font color is anything other than black (or default).
        Returns 1 if colored, 0 if black/default.
        """
        if cell.font.color is None:
            return 0

        if float(cell.font.color.tint) != 0.0:
            return 1

        if cell.font.color.theme is not None and cell.font.color.theme != 1:
            return 1

        if cell.font.color.rgb is not None:
            rgb_str = cell.font.color.rgb
            if isinstance(rgb_str, str) and len(rgb_str) == 8:
                # Extract hex
                font_color = f"#{rgb_str[2:].lower()}"
                if font_color != '#000000':
                    return 1
                else:
                    return 0
            else:
                return 0
        return 0

    # Suppress all UserWarnings related to xlsx
    warnings.filterwarnings("ignore", category=UserWarning)

    # 1) Allocate x_tok [max_rows, max_cols, pad_length] for input_ids
    x_tok = torch.zeros((max_rows, max_cols, pad_length), dtype=torch.long)

    # 2) Allocate x_masks [max_rows, max_cols, pad_length] for attention_mask
    x_masks = torch.zeros((max_rows, max_cols, pad_length), dtype=torch.long)

    # 3) Allocate y_tok [max_rows, max_cols, 17] for metadata
    y_tok = torch.zeros((max_rows, max_cols, 17), dtype=torch.long)

    # Define trivial sequence for default x_tok and x_masks
    # We'll build a Python list for the minimal input_ids
    minimal_ids = [tokenizer.cls_token_id, tokenizer.sep_token_id] + [tokenizer.pad_token_id]*(pad_length - 2)
    minimal_mask = [1, 1] + [0]*(pad_length - 2)

    # Initialize 3D tensors of size row x col x seq = 100x100x<pad_length> for input_ids and masks
    # Use above sequences to set them by default
    for r in range(max_rows):
        for c in range(max_cols):
            x_tok[r, c, :] = torch.tensor(minimal_ids, dtype=torch.long)
            x_masks[r, c, :] = torch.tensor(minimal_mask, dtype=torch.long)

    # Load workbook with openpyxl
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active  # use first sheet

    # Iterate over rows x cols
    for row, col in itertools.product(
        range(min(sheet.max_row, max_rows)),
        range(min(sheet.max_column, max_cols))
    ):
        # Convert to 1-based indexing
        cell_obj = sheet.cell(row=row+1, column=col+1)

        # Determine cell type from openpyxl
        cell_type = get_dataType(cell_obj.value, cell_obj.number_format)

        # Get cell value as string
        cell_value = get_value(cell_type, cell_obj)

        # If the cell is visually "filled" or not
        # For .xlsx, we check fill colors from openpyxl
        is_fill = 0
        if cell_obj.fill and cell_obj.fill.start_color:
            # Some logic for ignoring default color
            fill_idx = cell_obj.fill.start_color.index  # e.g. '00000000'
            if fill_idx not in ['00000000', '00FFFFFF', 0]:
                is_fill = 1

        # Alignment
        h_align, v_align = get_alignment(cell_obj)

        # Font info
        cell_fam = int(cell_obj.font.family) if cell_obj.font.family else 0
        cell_size = cell_obj.font.size or 0
        is_bold = int(cell_obj.font.bold)
        is_italic = int(cell_obj.font.italic)
        under_type = 1 if cell_obj.font.underline == 'single' else 2 if cell_obj.font.underline == 'double' else 0
        scr = 1 if cell_obj.font.vertAlign == 'superscript' else 2 if cell_obj.font.vertAlign == 'subscript' else 0
        is_colored = get_fontcol(cell_obj)

        # Borders
        top_b, bot_b, left_b, right_b = get_borders(cell_obj)

        # Merge info
        hmerge, vmerge = get_merge(row, col, sheet)

        # -------------- Tokenize with HF tokenizer --------------
        encoded_value = tokenizer.encode_plus(
            cell_value,
            max_length=pad_length,
            truncation=True,
            padding='max_length',
            return_tensors='pt'
        )

        # Store input_ids in x_tok, attention_mask in x_masks
        x_tok[row, col, :] = encoded_value['input_ids'][0]
        x_masks[row, col, :] = encoded_value['attention_mask'][0]

        # -------------- Fill y_tok with metadata --------------
        y_tok[row, col, 0] = cell_type           # cell type
        y_tok[row, col, 1] = is_fill            # fill
        y_tok[row, col, 2] = h_align            # halign
        y_tok[row, col, 3] = v_align            # valign
        y_tok[row, col, 4] = cell_fam           # font family
        y_tok[row, col, 5] = int(cell_size)     # font size
        y_tok[row, col, 6] = is_bold            # bold
        y_tok[row, col, 7] = is_italic          # italic
        y_tok[row, col, 8] = under_type         # underline
        y_tok[row, col, 9] = scr                # escapement
        y_tok[row, col, 10] = is_colored        # colour
        y_tok[row, col, 11] = top_b             # top border style
        y_tok[row, col, 12] = bot_b             # bottom border style
        y_tok[row, col, 13] = left_b            # left border style
        y_tok[row, col, 14] = right_b           # right border style
        y_tok[row, col, 15] = hmerge            # hmerge
        y_tok[row, col, 16] = vmerge            # vmerge

        # If empty/blank/unknown => further checks
        if y_tok[row, col, 0] in [0, 13, 14]:
            # Zero out font-related metadata
            y_tok[row, col, 4:11] = 0
            # If cell is "center across cells" => replicate from a neighbor
            if y_tok[row, col, 2] == 6:
                start_col = col
                while (
                    start_col >= 0 and
                    y_tok[row, start_col, 2] == 6 and
                    y_tok[row, start_col, 0] in (0, 13)
                ):
                    start_col -= 1
                # Copy x_tok, x_masks, y_tok from that column
                x_tok[row, col, :] = x_tok[row, start_col, :].clone()
                x_masks[row, col, :] = x_masks[row, start_col, :].clone()
                y_tok[row, col, :] = y_tok[row, start_col, :].clone()

    # -------------- Merge Ranges --------------
    for merged in sheet.merged_cells.ranges:
        start_cell, end_cell = str(merged).split(':')
        start_col_name, start_row_idx = coordinate_from_string(start_cell)
        end_col_name, end_row_idx = coordinate_from_string(end_cell)

        start_row_idx = start_row_idx - 1
        start_col_idx = column_index_from_string(start_col_name) - 1
        end_row_idx = min(end_row_idx - 1, max_rows - 1)
        end_col_idx = min(column_index_from_string(end_col_name) - 1, max_cols - 1)

        # Clip start row/col if needed
        if start_row_idx >= max_rows or start_col_idx >= max_cols:
            continue

        x_tok_start = x_tok[start_row_idx, start_col_idx, :].clone()
        x_masks_start = x_masks[start_row_idx, start_col_idx, :].clone()
        y_tok_start = y_tok[start_row_idx, start_col_idx, :].clone()

        for rr, cc in itertools.product(
            range(start_row_idx, end_row_idx + 1),
            range(start_col_idx, end_col_idx + 1)
        ):
            if rr < max_rows and cc < max_cols:
                x_tok[rr, cc, :] = x_tok_start
                x_masks[rr, cc, :] = x_masks_start
                y_tok[rr, cc, :15] = y_tok_start[:15]
                y_tok[rr, cc, 17:] = y_tok_start[17:]

    return x_tok, x_masks, y_tok

#######################################################################################
## CSV Function

def process_csv(file_path, tokenizer, max_rows=100, max_cols=100, pad_length=32):
    """
    Extracts metadata and tokenized values for each cell in a CSV file using a Hugging Face tokenizer.

    Args:
        file_path (str): The path to the CSV file.
        tokenizer: A Hugging Face tokenizer object (e.g., from transformers).
        max_rows (int, optional): The maximum number of rows to process. Defaults to 100.
        max_cols (int, optional): The maximum number of columns to process. Defaults to 100.
        pad_length (int, optional): The length to which each tokenized cell value is padded/truncated. Defaults to 32.

    Returns:
        Tuple[torch.LongTensor, torch.LongTensor, torch.LongTensor]: A tuple containing:
            - A (max_rows x max_cols x pad_length) PyTorch LongTensor with encoded tokenized values (input_ids).
            - A (max_rows x max_cols x pad_length) PyTorch LongTensor containing the attention masks.
            - A (max_rows x max_cols x 17) PyTorch LongTensor where each element is a tensor containing the cell's metadata.
    """

    def get_dataType(value: str) -> int:
        """
        Determines the type of the cell and handles necessary conversions.

        Int to Type Key Mapping:
            0  -> Empty cell
            1  -> Text cell
            2  -> Numeric type
            3  -> Integer subclass of number
            4  -> Float subclass of number
            5  -> Percentage subclass of number
            6  -> Currency subclass of number
            7  -> Scientific subclass of number
            8  -> Date type
            9  -> Time subclass of date
            10 -> Datetime subclass of date
            11 -> Boolean type with T or F value
            12 -> Error type corresponding to #REF!, #VALUE! etc. in Excel
            13 -> Blank cell with formatting
            14 -> Unknown type not in our keys

        Args:
            value (str): The value of the cell.

        Returns:
            int: The determined type of the cell (0-14).
        """

        # Check for blank cell
        if value == '':
            return 13

        # Check for date type
        date_patterns = [
            r"^\d{1,2}/\d{1,2}/\d{2,4}$",
            r"^\d{1,2}-\w{3}-\d{2,4}$",
            r"^\w{3,9}-\d{2}$",
            r"^\w{3,9}\s\d{1,2},\s\d{4}$",
            r"^\d{4}-\d{1,2}-\d{1,2}$"
        ]
        time_patterns = [
            r"^\d{1,2}:\d{2}(:\d{2})?(\.\d+)?\s?(AM|PM)?$",
            r"^\d{1,2}:\d{2}\s?(AM|PM)?$"
        ]
        datetime_patterns = [
            r"^\d{1,2}/\d{1,2}/\d{2,4}\s\d{1,2}:\d{2}(:\d{2})?\s?(AM|PM)?$",  # 1/20/00 12:33 AM, 1/20/00 12:33:00
            r"^\d{1,2}-\w{3}-\d{2,4}\s\d{1,2}:\d{2}(:\d{2})?\s?(AM|PM)?$",     # 1-Jan-00 12:33 AM
            r"^\d{1,2}/\d{1,2}/\d{4}\s\d{1,2}:\d{2}(:\d{2})?\s?(AM|PM)?$",     # 1/20/2000 12:33:00
        ]

        # Check for datetime first
        if any(re.fullmatch(dp, value) for dp in datetime_patterns):
            return 10

        # Check for date
        if any(re.fullmatch(dp, value.split()[0]) for dp in date_patterns):
            # Check for datetime
            if any(re.fullmatch(tp, value.split()[-1]) for tp in time_patterns):
                return 10
            else:
                return 8

        # Check for time
        if any(re.fullmatch(tp, value) for tp in time_patterns):
            return 9

        # Check for percentage
        if value.endswith('%') and re.match(r"^-?\d+(\.\d+)?%$", value):
            return 5

        # Check for currency
        if re.match(r"^[-\(\{\[]?\s*[\$£€₹¥]?\s*\d+(\.\d{2})?\s*[\)\}\]]?$", value):
            return 6

        # Check for scientific notation
        if re.match(r"^[+-]?\d+(\.\d+)?[eE][+-]?\d+$", value):
            return 7

        # Check for float type
        if re.match(r"^[+-]?\d+\.\d+$", value):
            return 4

        # Check for integer type
        if re.match(r"^[+-]?\d+$", value):
            return 3

        # Check for boolean
        if value.lower() in {'true', 'false'}:
            return 11

        # Check for error type
        if value in {
            "#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!"
        }:
            return 12

        # Check for str
        if isinstance(value, str):
            return 1

        # Default to unk if none of the above
        return 14

    # Detect the delimiter from the file
    with open(file_path, mode='r', newline='', encoding='utf-8') as csvfile:
        delimiter = csv.Sniffer().sniff(csvfile.read()).delimiter

    # Initialize a (max_rows x max_cols x pad_length) LongTensor filled with zeros for tokenized data
    x_tok = torch.zeros((max_rows, max_cols, pad_length), dtype=torch.long)

    # Initialize a matching LongTensor for attention masks
    x_masks = torch.zeros((max_rows, max_cols, pad_length), dtype=torch.long)

    # Initialize a (max_rows x max_cols x 17) LongTensor filled with zeros for metadata
    y_tok = torch.zeros((max_rows, max_cols, 17), dtype=torch.long)

    # Define trivial sequence for default x_tok and x_masks
    # We'll build a Python list for the minimal input_ids
    minimal_ids = [tokenizer.cls_token_id, tokenizer.sep_token_id] + [tokenizer.pad_token_id]*(pad_length - 2)
    minimal_mask = [1, 1] + [0]*(pad_length - 2)

    # Initialize 3D tensors of size row x col x seq = 100x100x<pad_length> for input_ids and masks
    # Use above sequences to set them by default
    for r in range(max_rows):
        for c in range(max_cols):
            x_tok[r, c, :] = torch.tensor(minimal_ids, dtype=torch.long)
            x_masks[r, c, :] = torch.tensor(minimal_mask, dtype=torch.long)


    # Open the file and read it as CSV using the detected delimiter
    with open(file_path, mode='r', newline='', encoding='utf-8') as file:
        reader = csv.reader(file, delimiter=delimiter)

        # Iterate through each row in the CSV file
        for row_index, row_data in enumerate(reader):

            if row_index >= max_rows:
                break

            # Iterate through each column in the row
            for col_index, cell_value in enumerate(row_data):

                if col_index >= max_cols:
                    break

                # Only process cells within the max_rows x max_cols limit
                if row_index < max_rows and col_index < max_cols:

                    # Determine the type of the cell using the get_dataType function
                    cell_val_str = cell_value.strip()
                    cell_type = get_dataType(cell_val_str)

                    # Use the Hugging Face tokenizer (replacing the old tokenize_pad)
                    encoded_value = tokenizer.encode_plus(
                        cell_val_str,
                        max_length=pad_length,
                        truncation=True,
                        padding='max_length',
                        return_tensors='pt'  # shape => [1, pad_length]
                    )

                    # Store input_ids in x_tok
                    x_tok[row_index, col_index, :] = encoded_value['input_ids'][0]

                    # Store attention_mask in x_masks
                    x_masks[row_index, col_index, :] = encoded_value['attention_mask'][0]

                    # If blank cell with no type info, then set to 0s
                    if cell_type == 13:
                        y_tok[row_index, col_index, :] = torch.zeros(17, dtype=torch.long)

                    else:
                        # Store the metadata in y_tok
                        y_tok[row_index, col_index, 0] = cell_type
                        y_tok[row_index, col_index, 1] = 0
                        y_tok[row_index, col_index, 2] = 0
                        y_tok[row_index, col_index, 3] = 2
                        y_tok[row_index, col_index, 4] = 0
                        y_tok[row_index, col_index, 5] = 12
                        y_tok[row_index, col_index, 6] = 0
                        y_tok[row_index, col_index, 7] = 0
                        y_tok[row_index, col_index, 8] = 0
                        y_tok[row_index, col_index, 9] = 0
                        y_tok[row_index, col_index, 10] = 0
                        y_tok[row_index, col_index, 11] = 0
                        y_tok[row_index, col_index, 12] = 0
                        y_tok[row_index, col_index, 13] = 0
                        y_tok[row_index, col_index, 14] = 0
                        y_tok[row_index, col_index, 15] = 0
                        y_tok[row_index, col_index, 16] = 0

    # Return the 3D LongTensors for (input_ids), (attention_mask), and metadata
    return x_tok, x_masks, y_tok

#######################################################################################
## Process Spreadsheet Function

def process_spreadsheet(
    file_path: str,
    tokenizer,                 # A Hugging Face ModernBert tokenizer
    max_rows=100,
    max_cols=100,
    pad_length=32
):
    """
    Processes a spreadsheet file (.xls, .xlsx, .csv) and extracts both the tokenized train tensors 
    and the metadata tensor, using a Hugging Face tokenizer for cell content.

    Args:
        file_path (str): The path to the spreadsheet file.
        tokenizer: A Hugging Face tokenizer object (replacing the old vocabulary approach).
        max_rows (int, optional): The maximum number of rows to process. Defaults to 100.
        max_cols (int, optional): The maximum number of columns to process. Defaults to 100.
        pad_length (int, optional): The length to which the tokenized cell value is padded/truncated. Defaults to 32.

    Returns:
        A tuple containing:
            - x_tok (torch.LongTensor): A (max_rows x max_cols x pad_length) tensor with input_ids.
            - x_masks (torch.LongTensor): A (max_rows x max_cols x pad_length) tensor with attention masks.
            - y_tok (torch.LongTensor): A (max_rows x max_cols x 17) tensor containing metadata for each cell.

    Raises:
        Exception: If an unsupported file format is given or any other error occurs during processing.
    """

    try:
        file_extension = file_path.split('.')[-1].lower()

        if file_extension not in ['xls', 'xlsx', 'csv']:
            raise ValueError(f"Unsupported file format: {file_extension} for {file_path}")

        if file_extension == 'xls':
            x_tok, x_masks, y_tok = process_xls(
                file_path, tokenizer, max_rows, max_cols, pad_length
            )
        elif file_extension == 'xlsx':
            x_tok, x_masks, y_tok = process_xlsx(
                file_path, tokenizer, max_rows, max_cols, pad_length
            )
        else:  # 'csv'
            x_tok, x_masks, y_tok = process_csv(
                file_path, tokenizer, max_rows, max_cols, pad_length
            )

        return x_tok, x_masks, y_tok

    except Exception as e:
        raise e
