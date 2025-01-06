# Imports for current file
import importlib
from utils import selfutil
importlib.reload(selfutil)
# exec(open("./selfutil.py").read())

# Import parsing libraries
import xlrd
from xlrd.xldate import xldate_as_tuple, XLDateAmbiguous, XLDateError
import openpyxl
import csv
import re

# Import general usecase libraries
import pandas as pd
import torch
from datetime import datetime, time
import warnings
import itertools
from typing import Optional, Tuple
import sys


# Func to tokenize/pad for saffu 
def saffu_tokenize_pad(text, tokenizer, pad_length):
    """
    Tokenizes a given string using the provided tokenizer, adds start and end tokens, and pads/truncates to the specified pad length.

    Args:
        text (str): The text to be tokenized.
        tokenizer (SAFFUTokenizer): The tokenizer to use for tokenization.
        pad_length (int): The length to which the tokenized list should be padded or truncated.

    Returns:
        list: A list of tokens padded or truncated to the specified length.
    """

    # Define condition
    #cond = len(text) < pad_length

    # Tokenize the input text
    tokenized_text = tokenizer._tokenize(text)

    # cond = len(tokenized_text) >= 20

    # if cond:
    #     print(f'Initial Text ({len(text)}):\n{text}\n')
    #     print(f'Tokenized Text ({len(tokenized_text)}):\n{tokenized_text}\n')

    # Add <eod> at the end of the text
    tokenized_text = tokenized_text + [tokenizer._eod]
    
    # if cond:
    #     print(f'EOD Front Add ({len(tokenized_text)}):\n{tokenized_text}\n')

    # If tokenized text is less than the pad length provided then add pad token to desired length
    if len(tokenized_text) < pad_length:
        tokenized_text += [tokenizer._pad] * (pad_length - len(tokenized_text))
    
    # Otherwise truncate to desired pad_length 
    else:
        tokenized_text = tokenized_text[:pad_length]

    # if cond:
    #     print(f'PAD Front Add ({len(tokenized_text)}):\n{tokenized_text}\n')
        
    # Now add <sod> token at the beginning of text taking length to pad_length + 1
    tokenized_text = [tokenizer._sod] + tokenized_text

    # if cond:
    #     print(f'SOD Back Add ({len(tokenized_text)}):\n{tokenized_text}\n')

    # Then add <pad> tokens as per config._r at the start taking length to pad_length + 1 + config._r 
    tokenized_text = [tokenizer._pad]*tokenizer.config._r + tokenized_text

    # if cond:
    #     print(f'PAD Back Add ({len(tokenized_text)}):\n{tokenized_text}\n')
    
    # Encode the tokenized text
    encoded_text = [tokenizer._convert_token_to_id(token) for token in tokenized_text]

    # if cond:
    #     print(f'Encoded Text ({len(encoded_text)}):\n{encoded_text}\n')
    #     sys.exit(0)

    # Return the final tokenized, padded text
    return encoded_text


# Func to process xlsx file
def process_xlsx(file_path, tokenizer, max_rows, max_cols, pad_length):

    def get_dataType(value, number_format):
        """
        Determines the type of the cell and handles necessary conversions.

        Int to Type Key Mapping:
            0  -> Empty cell
            1  -> Text cell
            2  -> Numeric type
            3  -> Integer subclass of number
            4  -> Float subclass of number
            5  -> Percentage subclass of number
            6  -> Currency subclass of number
            7  -> Scientific subclass of number
            8  -> Date type
            9  -> Time subclass of date
            10 -> Datetime subclass of date
            11 -> Boolean type with T or F value
            12 -> Error type corresponding to #REF!, #VALUE! etc. in excel
            13 -> Blank cell with formatting
            14 -> Unknown type not in our keys

        Args:
            value: The value of the cell.
            number_format (str): The number format of the cell.

        Returns:
            int: The determined type of the cell (0-14).
        """

        # Empty cell with no value return blank for this
        if value is None:
            return 13

        # Boolean cell
        elif isinstance(value, bool):
            return 11

        # Error cell based on certain error values in Excel
        elif any(error in str(value) for error in {"#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}):
            return 12

        # Currency format cell
        elif "#,##0" in number_format:
            return 6

        # Percentage format cell
        elif number_format.endswith("%"):
            return 5

        # Check for scientific notation format
        elif "E+" in number_format or "E-" in number_format:
            return 7

        # If float type
        elif isinstance(value, float):

            # Check if it's an integer subclass
            if value.is_integer():
                return 3

            # Otherwise, it's a float
            return 4

        # If integer type
        elif isinstance(value, int):
            return 3

        # Time type cell
        elif isinstance(value, time):
            return 9

        # Datetime type cell
        elif isinstance(value, datetime):

            # Check if the time part is exactly midnight
            if value.time() == time(0, 0):
                # If the datetime is formatted as a date, return it as a date type
                if "d" in number_format.lower() and "h" not in number_format.lower():
                    return 8
                else:
                    # If there is significant time information, return it as datetime
                    return 10
            else:
                # If the time part is not midnight, it's definitely a datetime
                return 10

        # Blank cell with formatting
        elif value == '':
            return 13

        # Text cell
        elif isinstance(value, str):
            return 1

        # Return 14 for any unknown type
        return 14

    def get_value(cell_type, cell):
        # If empty cell then return an empty string
        if cell_type == 0:
            return ''

        # Text cell
        elif cell_type == 1:
            return str(cell.value)

        # Numeric cell default type
        elif cell_type == 2:
            return str(cell.value)

        # Integer subclass of number, remove proceeding 0s and return str
        elif cell_type == 3:
            return str(cell.value)

        # Decimal subclass then convert to float and return str
        elif cell_type == 4:
            return str(cell.value)

        # Percentage subclass then convert to value add % sign and return the string
        elif cell_type == 5:
            return str(cell.value * 100) + '%'

        # Currency subclass then convert to value add currency symbol and return the string
        elif cell_type == 6:
            # Check if cell value already is a string so it will have the symbol
            if isinstance(cell.value, str):
                return str(cell.value)

            # Check if euro sign in format string because that is separate
            if '$€' in cell.number_format:
                return '€' + str(float(cell.value))

            # In other case for usd values we gotta add the $ symbol
            return '$' + str(float(cell.value))

        # If scientific subclass then convert to value add 'E' sign and return the string
        elif cell_type == 7:
            return f"{cell.value:.2e}"

        # If date subclass then use the df to return value
        elif cell_type == 8 or cell_type == 9 or cell_type == 10:
            return str(cell.value)

        # If boolean cell then use df to get string
        elif cell_type == 11:
            return str(cell.value)

        # If error cell then use df to get string
        elif cell_type == 12:
            return str(cell.value)

        # If blank cell (Empty cell with formatting) then return an empty string
        elif cell_type == 13:
            return ''

        # Return empty string for default case
        return ''

    def get_alignment(cell):

        # Define dict to store the alignment keys mapped to xlrd
        openpyxl_to_xlrd = {
            'horizontal': {'general': 0, 'left': 1, 'center': 2, 'right': 3, 'fill': 4, 'justify': 5, 'centerContinuous': 6, 'distributed': 7},
            'vertical': {'top': 0, 'center': 1, 'bottom': 2, 'justify': 3, 'distributed': 4},
        }

        # Get the horizontal and vertical alignment else default if key not found
        horiz = openpyxl_to_xlrd['horizontal'].get(cell.alignment.horizontal, 0)
        vert = openpyxl_to_xlrd['vertical'].get(cell.alignment.vertical, 2)

        # Concatenate the keys and return as int
        return horiz, vert

    def get_borders(cell):

        # Define dict to store the border keys mapped to xlrd
        openpyxl_to_xlrd = {
            'none': 0, 'thin': 1, 'medium': 2, 'dashed': 3, 'dotted': 4, 'thick': 5, 'double': 6, 'hair': 7, 'mediumDashed': 8, 'dashDot': 9, 'mediumDashDot': 10, 'dashDotDot': 11,
            'mediumDashDotDot': 12, 'slantDashDot': 13
        }

        return openpyxl_to_xlrd.get(cell.border.top.style or 'none'), openpyxl_to_xlrd.get(cell.border.bottom.style or 'none'), openpyxl_to_xlrd.get(
            cell.border.left.style or 'none'), openpyxl_to_xlrd.get(cell.border.right.style or 'none')

    def get_merge(row, col, sheet):
        """
        Determines the merge status of a cell in terms of its vertical and horizontal position within a merged range.

        Int to Merge Status Mapping:
            1 -> Outside any merged range ('O')
            2 -> Beginning of a merged range ('B')
            3 -> Inside a merged range but not at the beginning ('I')

        Args:
            row (int): The row index of the cell.
            col (int): The column index of the cell.
            sheet (openpyxl.worksheet.worksheet.Worksheet): The sheet object containing the merged cells information.

        Returns:
            int: A 2-digit integer where the first digit represents the vertical position
                 and the second digit represents the horizontal position within a merged range.
        """

        # Initialize tags as 1 (outside) by default
        vertical_tag = 0
        horizontal_tag = 0

        # Convert 0-based indexing to 1-based indexing for openpyxl
        row_1_based = row + 1
        col_1_based = col + 1
        cell_coord = openpyxl.utils.get_column_letter(col_1_based) + str(row_1_based)

        # Iterate over the merged cells in the sheet
        for merged in sheet.merged_cells.ranges:

            # Check if the current cell is within a merged range
            if cell_coord in merged:

                # Determine vertical tag
                if row_1_based == merged.min_row:
                    vertical_tag = 1  # Beginning of vertical range
                else:
                    vertical_tag = 2  # Inside vertical range

                # Determine horizontal tag
                if col_1_based == merged.min_col:
                    horizontal_tag = 1  # Beginning of horizontal range
                else:
                    horizontal_tag = 2  # Inside horizontal range

                # Break the loop as the cell is part of a merged range
                break

        # Return the combined vertical and horizontal tags as a 2-digit integer
        return horizontal_tag, vertical_tag

    def get_fontcol(cell):

        # Check if total none then return 0
        if cell.font.color is None:
            return 0

        # Check tint first of all
        if float(cell.font.color.tint) != 0.0:
            return 1

        # Check theme
        if cell.font.color.theme is not None and cell.font.color.theme != 1:
            return 1

        # Check if rgb exists and is not None
        if cell.font.color.rgb is not None:

            # Ensure the RGB value is 8 characters long (including opacity) and starts with 'FF' for full opacity
            if isinstance(cell.font.color.rgb, str) and len(cell.font.color.rgb) == 8:

                # Extract the last 6 characters which represent the color and check if it's not black
                font_color = f"#{cell.font.color.rgb[2:].lower()}"

                # If color is not black return 1 else 0
                if font_color != '#000000':
                    return 1  # Font is colored
                else:
                    return 0  # Font is black (default)
            else:
                return 0  # RGB value doesn't match expected format

        # If rgb is None
        else:
            return 0

    # Suppress warnings related to headers and footers
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.worksheet.header_footer")
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.worksheet._reader")
    warnings.filterwarnings("ignore", category=UserWarning)

    # Initialize 3D tensor of size rows x cols x pad_length = 100 x 100 x (32 + _r + 1) = 100x100x35 necessary for tokenize_pad also
    x_tok = torch.zeros((max_rows, max_cols, pad_length + tokenizer.config._r + 1), dtype=torch.long) ########################################### changes happened 3 x here!!

    # Initialize y_tok tensor to store metadata for each cell of size row x col x 6
    y_tok = torch.zeros((max_rows, max_cols, 17), dtype=torch.long)

    # Load the workbook and access the active sheet
    workbook = openpyxl.load_workbook(file_path, data_only=True)

    # Retrieve the sheet
    sheet = workbook.active

    # Iterate over each combination of rows and columns in the sheet, up to the 100th row and column
    for row, col in itertools.product(range(min(sheet.max_row, max_rows)), range(min(sheet.max_column, max_cols))):

        # Adjust to 1-based indexing
        cell = sheet.cell(row=row + 1, column=col + 1)

        # Determine the cell type
        cell_type = get_dataType(cell.value, cell.number_format)

        # Determine the value of the cell
        cell_value = get_value(cell_type, cell)


        # Adjusted logic for determining if the cell is filled
        is_fill = 0 if (cell.fill.start_color.index if cell.fill and cell.fill.start_color else None) in ['00000000', '00FFFFFF', 0] else 1

        # Get alignment of the cell
        halign, valign = get_alignment(cell)

        # Determine font family if None then default to 0
        cell_fam = int(cell.font.family) if cell.font.family else 0

        # Determine font size
        cell_size = cell.font.size

        # Determine cell style
        is_bold = int(cell.font.bold)
        is_italic = int(cell.font.italic)
        cell_underline = 1 if cell.font.underline == 'single' else 2 if cell.font.underline == 'double' else 0
        cell_scr = 1 if cell.font.vertAlign == 'superscript' else 2 if cell.font.vertAlign == 'subscript' else 0

        # is_col = 0 if font_color == '#000000' else 1
        is_col = get_fontcol(cell)

        # Determine borders
        btop, bbot, bleft, bright = get_borders(cell)

        # Get merged range
        hmerge, vmerge = get_merge(row, col, sheet)

        # Set value to appropriate location in x_tok
        x_tok[row, col, :] = torch.tensor(saffu_tokenize_pad(cell_value, tokenizer, pad_length), dtype=torch.long)

        # Check if empty cell or blank cell with no metadata
        if cell_type == 0 or (cell_type == 13 and  # Blank cell type
                              is_fill == 0 and  # No fill
                              halign == 0 and valign == 2 and  # Default alignment
                              is_bold == 0 and  # No bold
                              is_italic == 0 and  # No italic
                              cell_underline == 0 and  # No underline
                              cell_scr == 0 and  # No super or subscript
                              is_col == 0 and  # No colour
                              btop == 0 and  # No top border
                              bbot == 0 and  # No bottom border
                              bleft == 0 and  # No left border
                              bright == 0 and  # No right border
                              hmerge == 0 and vmerge == 0):

            # Set to 0 tensor
            y_tok[row, col, :] = torch.zeros(17, dtype=torch.long)

        # In normal case
        else:
            y_tok[row, col, 0] = cell_type
            y_tok[row, col, 1] = is_fill  # Store 'fill'
            y_tok[row, col, 2] = halign  # Store 'halign'
            y_tok[row, col, 3] = valign  # Store 'valign'
            y_tok[row, col, 4] = cell_fam  # Store 'font family'
            y_tok[row, col, 5] = cell.font.size  # Store 'font size'
            y_tok[row, col, 6] = is_bold  # Store 'bold'
            y_tok[row, col, 7] = is_italic  # Store 'italic'
            y_tok[row, col, 8] = cell_underline  # Store 'underline'
            y_tok[row, col, 9] = cell_scr  # Store 'escapement'
            y_tok[row, col, 10] = is_col  # Store 'colour'
            y_tok[row, col, 11] = btop  # Store top border style
            y_tok[row, col, 12] = bbot  # Store bottom border style
            y_tok[row, col, 13] = bleft  # Store 'left border style'
            y_tok[row, col, 14] = bright  # Store 'right border style'
            y_tok[row, col, 15] = hmerge
            y_tok[row, col, 16] = vmerge  # Store 'hmerge' and 'vmerge'

    # Return the 2D NumPy array containing the metadata for each cell
    return x_tok, y_tok



    
    
# Func to process xls file
def process_xls(file_path, tokenizer, max_rows, max_cols, pad_length):
    """
    Extracts metadata for each cell in the first sheet of the given spreadsheet and tokenizes the cell values.

    Args:
        file_path (str): The path to the spreadsheet file.
        vocab (dict): A dictionary mapping tokens to their indices for encoding.
        max_rows (int): The maximum number of rows to process.
        max_cols (int): The maximum number of columns to process.
        pad_length (int): The length to which the tokens should be padded.

    Returns:
        x_tok (torch.Tensor): A 3D tensor of size (max_rows, max_cols, pad_length) containing tokenized cell values.
        y_tok (torch.Tensor): A 3D tensor of size (max_rows, max_cols, 7) containing metadata for each cell:
            - Index 0: Cell type (int)
            - Index 1: Fill (int)
            - Index 2: Alignment (int)
            - Index 3: Font family (int)
            - Index 4: Style (int)
            - Index 5: Borders (int)
            - Index 6: Merge status (int)
    """

    def get_fill(xf_record, workbook):
        """
        Determines whether the cell is filled with a background color.

        Args:
            xf_record (xlrd.formatting.XFRecord): The XFRecord object containing formatting information for the cell.
            workbook (xlrd.Book): The workbook object containing the cell.

        Returns:
            int: 1 if the cell is filled with a color other than white, 0 otherwise.
        """

        # Retrieve the color of the cell's background using the pattern_colour_index
        bg_color_rgb = workbook.colour_map.get(xf_record.background.pattern_colour_index)

        # Return 0 if the cell is not filled or filled with white; otherwise, return 1
        return 0 if bg_color_rgb in [None, (255, 255, 255)] else 1

    def get_merge(row, col, sheet):
        """
        Determines the merge status of a cell in terms of its vertical and horizontal position within a merged range.

        Int to Merge Status Mapping:
            1 -> Outside any merged range ('O')
            2 -> Beginning of a merged range ('B')
            3 -> Inside a merged range but not at the beginning ('I')

        Args:
            row (int): The row index of the cell.
            col (int): The column index of the cell.
            sheet (xlrd.sheet.Sheet): The sheet object containing the merged cells information.

        Returns:
            int: A 2-digit integer where the first digit represents the vertical position
                 and the second digit represents the horizontal position within a merged range.
        """

        # Initialize tags as 0 (outside) by default
        vertical_tag = 0
        horizontal_tag = 0

        # Iterate over the merged cells in the sheet
        for start_row, end_row, start_col, end_col in sheet.merged_cells:

            # Check if the current cell is within a merged range
            if start_row <= row < end_row and start_col <= col < end_col:

                # Determine vertical tag
                if row == start_row:
                    vertical_tag = 1  # Beginning of vertical range
                else:
                    vertical_tag = 2  # Inside vertical range

                # Determine horizontal tag
                if col == start_col:
                    horizontal_tag = 1  # Beginning of horizontal range
                else:
                    horizontal_tag = 2  # Inside horizontal range

                # Break the loop as the cell is part of a merged range
                break

        # Return the combined vertical and horizontal tags
        return horizontal_tag, vertical_tag

    def get_data(cell, workbook, df, row, col):
        """
        Determines the type of the cell, handles necessary conversions, and retrieves the cell value.

        Args:
            cell (xlrd.sheet.Cell): The cell object to determine the type and retrieve the value.
            workbook (xlrd.Book): The workbook object containing the cell.
            df (pandas.DataFrame): DataFrame representation of the sheet for value extraction.
            row (int): The row index of the cell.
            col (int): The column index of the cell.

        Returns:
            tuple: A tuple containing the determined type of the cell (int) and the cell value (str).
        """
        xlrd_error_map = {
            0: "#NULL!", 7: "#DIV/0!", 15: "#VALUE!", 23: "#REF!", 29: "#NAME?", 36: "#NUM!", 42: "#N/A", 43: "#GETTING_DATA",
        }

        # Empty cell with no formatting
        if cell.ctype == 0:
            return 0, ''

        # If text cell then check further to get exact text type
        elif cell.ctype == 1:
            format_str = workbook.format_map[workbook.xf_list[cell.xf_index].format_key].format_str

            if '#,##0' in format_str:
                return 6, str(cell.value)  # Currency type

            return 1, str(cell.value)  # Text type

        # If numeric cell then check further to get exact numeric type
        elif cell.ctype == 2:
            format_str = workbook.format_map[workbook.xf_list[cell.xf_index].format_key].format_str

            if format_str == '0.00E+00':
                return 7, f"{cell.value:.2e}"  # Scientific subclass of number

            elif '#,##0' in format_str:

                if '$€' in format_str:
                    return 6, '€' + str(float(cell.value))

                return 6, '$' + str(float(cell.value))

            elif '%' in format_str:
                return 5, str(cell.value * 100) + '%'  # Percentage type

            elif isinstance(cell.value, float):
                if cell.value.is_integer():
                    return 3, str(df.iat[row, col])  # Integer type
                return 4, str(df.iat[row, col])  # Float type

            return 2, str(df.iat[row, col])  # General numeric type

        # If date cell then check further to get exact date type
        elif cell.ctype == 3:
            try:
                date_tuple = xldate_as_tuple(cell.value, workbook.datemode)

                if date_tuple[:3] == (0, 0, 0):
                    return 9, str(df.iat[row, col])  # Time type

                elif date_tuple[3:] == (0, 0, 0):
                    return 8, str(df.iat[row, col])  # Date type

                return 10, str(df.iat[row, col])  # Datetime type

            except (XLDateError, XLDateAmbiguous):
                return 10, str(df.iat[row, col])  # Datetime type

        # Boolean cell
        elif cell.ctype == 4:
            return 11, str(df.iat[row, col])

        # Error cell
        elif cell.ctype == 5:
            return 12, xlrd_error_map.get(cell.value, f"#ERR{cell.value}")

        # Blank cell (Empty cell with formatting)
        elif cell.ctype == 6:
            return 13, ''

        # Unknown type not in our keys return by default
        return 14, ''

    # Initialize 3D tensor of size rows x cols x pad_length = 100 x 100 x (32 + _r + 1) = 100x100x35 necessary for tokenize_pad also
    x_tok = torch.zeros((max_rows, max_cols, pad_length + tokenizer.config._r + 1), dtype=torch.long)

    # Initialize y_tok tensor to store metadata for each cell of size row x col x 17
    y_tok = torch.zeros((max_rows, max_cols, 17), dtype=torch.long)

    # Open the .xls file with formatting_info=True to get the metadata
    workbook = xlrd.open_workbook(filename=file_path, formatting_info=True)

    # Access the first sheet in the workbook
    sheet = workbook.sheet_by_index(0)

    # Get the df for some parsing
    df_read = pd.read_excel(file_path, header=None, dtype=str, na_values=' ', keep_default_na=False, engine='xlrd')

    # Convert to the size of df required
    df = df_read.reindex(index=range(max_rows), columns=range(max_cols), fill_value='')

    # Iterate over each combination of rows and columns in the sheet, up to the 100th row and column
    for row, col in itertools.product(range(min(sheet.nrows, max_rows)), range(min(sheet.ncols, max_cols))):

        # Retrieve the cell object from the sheet
        cell = sheet.cell(row, col)

        # Access the XFRecord object using the cell's XF index
        xf_record = workbook.xf_list[cell.xf_index]

        # Get cell train
        cell_type, cell_value = get_data(cell, workbook, df, row, col)

        # Set the tokens for current cell in x_tok
        x_tok[row, col, :] = torch.tensor(saffu_tokenize_pad(cell_value, tokenizer, pad_length), dtype=torch.long)

        # Determine the fill of the cell
        y_tok[row, col, 0] = cell_type
        y_tok[row, col, 1] = get_fill(xf_record, workbook)  # Store 'fill'
        y_tok[row, col, 2] = xf_record.alignment.hor_align  # Store 'halign'
        y_tok[row, col, 3] = xf_record.alignment.vert_align  # Store 'valign'
        y_tok[row, col, 4] = workbook.font_list[xf_record.font_index].family  # Store 'font family'
        y_tok[row, col, 5] = workbook.font_list[xf_record.font_index].height // 20  # Store 'font size'
        y_tok[row, col, 6] = workbook.font_list[xf_record.font_index].bold  # Store 'bold'
        y_tok[row, col, 7] = workbook.font_list[xf_record.font_index].italic  # Store 'italic'
        y_tok[row, col, 8] = workbook.font_list[xf_record.font_index].underline_type  # Store 'underline'
        y_tok[row, col, 9] = workbook.font_list[xf_record.font_index].escapement  # Store 'escapement'
        y_tok[row, col, 10] = 1 if (rgb := workbook.colour_map.get(workbook.font_list[xf_record.font_index].colour_index)) and rgb != (0, 0, 0) else 0  # Store 'colour'
        y_tok[row, col, 11] = xf_record.border.top_line_style  # Store top border style
        y_tok[row, col, 12] = xf_record.border.bottom_line_style  # Store bottom border style
        y_tok[row, col, 13] = xf_record.border.left_line_style  # Store 'left border style'
        y_tok[row, col, 14] = xf_record.border.right_line_style  # Store 'right border style'
        y_tok[row, col, 15], y_tok[row, col, 16] = get_merge(row, col, sheet)  # Store 'hmerge' and 'vmerge'

        # Check if empty cell or blank cell with no metadata
        if y_tok[row, col, 0] == 0 or (
                y_tok[row, col, 0] == 13 and y_tok[row, col, 3] == 2 and all(y_tok[row, col, pos] == 0 for pos in [i for i in range(1, 17) if i not in [3, 4, 5]])
        ):
            # Set to 0 tensor
            y_tok[row, col, :] = torch.zeros(17, dtype=torch.long)

    # Return the 100x100 2D NumPy array containing the metadata for each cell
    return x_tok, y_tok

def process_csv(file_path, tokenizer, max_rows=100, max_cols=100, pad_length=32):
    """
    Extracts metadata and tokenized values for each cell in a CSV file.

    Args:
        file_path (str): The path to the CSV file.
        vocab: Vocabulary object for encoding tokens.
        max_rows (int, optional): The maximum number of rows to process. Defaults to 100.
        max_cols (int, optional): The maximum number of columns to process. Defaults to 100.
        pad_length (int, optional): The length to which each tokenized cell value is padded. Defaults to 32.

    Returns:
        Tuple[torch.LongTensor, torch.LongTensor]: A tuple containing:
            - A 100x100x32 PyTorch LongTensor with encoded tokenized values.
            - A 100x100x15 PyTorch LongTensor where each element is a tensor containing the cell's metadata.
    """

    def get_dataType(value: str) -> int:
        """
        Determines the type of the cell and handles necessary conversions.

        Int to Type Key Mapping:
            0  -> Empty cell
            1  -> Text cell
            2  -> Numeric type
            3  -> Integer subclass of number
            4  -> Float subclass of number
            5  -> Percentage subclass of number
            6  -> Currency subclass of number
            7  -> Scientific subclass of number
            8  -> Date type
            9  -> Time subclass of date
            10 -> Datetime subclass of date
            11 -> Boolean type with T or F value
            12 -> Error type corresponding to #REF!, #VALUE! etc. in Excel
            13 -> Blank cell with formatting
            14 -> Unknown type not in our keys

        Args:
            value (str): The value of the cell.

        Returns:
            int: The determined type of the cell (0-14).
        """

        # Check for blank cell
        if value == '':
            return 13

        # Check for date type
        date_patterns = [
            r"^\d{1,2}/\d{1,2}/\d{2,4}$",
            r"^\d{1,2}-\w{3}-\d{2,4}$",
            r"^\w{3,9}-\d{2}$",
            r"^\w{3,9}\s\d{1,2},\s\d{4}$",
            r"^\d{4}-\d{1,2}-\d{1,2}$"
        ]
        time_patterns = [
            r"^\d{1,2}:\d{2}(:\d{2})?(\.\d+)?\s?(AM|PM)?$",
            r"^\d{1,2}:\d{2}\s?(AM|PM)?$"
        ]
        datetime_patterns = [
            r"^\d{1,2}/\d{1,2}/\d{2,4}\s\d{1,2}:\d{2}(:\d{2})?\s?(AM|PM)?$",  # 1/20/00 12:33 AM, 1/20/00 12:33:00
            r"^\d{1,2}-\w{3}-\d{2,4}\s\d{1,2}:\d{2}(:\d{2})?\s?(AM|PM)?$",  # 1-Jan-00 12:33 AM
            r"^\d{1,2}/\d{1,2}/\d{4}\s\d{1,2}:\d{2}(:\d{2})?\s?(AM|PM)?$",  # 1/20/2000 12:33:00
        ]

        # Check for datetime first
        if any(re.fullmatch(dp, value) for dp in datetime_patterns):
            return 10

        # Check for date
        if any(re.fullmatch(dp, value.split()[0]) for dp in date_patterns):
            # Check for datetime
            if any(re.fullmatch(tp, value.split()[-1]) for tp in time_patterns):
                return 10
            else:
                return 8

        # Check for time
        if any(re.fullmatch(tp, value) for tp in time_patterns):
            return 9

        # Check for percentage
        if value.endswith('%') and re.match(r"^-?\d+(\.\d+)?%$", value):
            return 5


        # Check for scientific notation
        if re.match(r"^[+-]?\d+(\.\d+)?[eE][+-]?\d+$", value):
            return 7

        # Check for float type
        if re.match(r"^[+-]?\d+\.\d+$", value):
            return 4

        # Check for integer type
        if re.match(r"^[+-]?\d+$", value):
            return 3

        # Check for currency
        if re.match(r"^[-\(\{\[]?\s*[\$£€₹¥]?\s*\d+(\.\d{2})?\s*[\)\}\]]?$", value):
            return 6

        # Check for boolean
        if value.lower() in {'true', 'false'}:
            return 11

        # Check for error type
        if value in {"#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}:
            return 12

        # Check for str
        if isinstance(value, str):
            return 1

        # Default to unk if none of the above
        return 14

    # Get the delimiter from the file
    with open(file_path, mode='r', newline='', encoding='utf-8') as csvfile:
        delimiter = csv.Sniffer().sniff(csvfile.read()).delimiter

    # Initialize 3D tensor of size rows x cols x pad_length = 100 x 100 x (32 + _r + 1) = 100x100x35 necessary for tokenize_pad also
    x_tok = torch.zeros((max_rows, max_cols, pad_length + tokenizer.config._r + 1), dtype=torch.long)

    # Initialize a 100x100x15 LongTensor filled with zeros for metadata
    y_tok = torch.zeros((max_rows, max_cols, 17), dtype=torch.long)

    # Open the file and read it as a CSV using the detected delimiter
    with open(file_path, mode='r', newline='', encoding='utf-8') as file:
        reader = csv.reader(file, delimiter=delimiter)

        # Iterate through each row in the CSV file
        for row_index, row in enumerate(reader):

            if row_index >= max_rows:
                break

            # Iterate through each column in the row
            for col_index, cell_value in enumerate(row):

                if col_index >= max_cols:
                    break

                # Only process cells within the 100x100 limit
                if row_index < max_rows and col_index < max_cols:

                    # Determine the type of the cell using the get_dataType function
                    cell_type = get_dataType(cell_value.strip())

                    # Store the tokenized and padded train in the tensor
                    x_tok[row_index, col_index, :] = torch.tensor(saffu_tokenize_pad(cell_value, tokenizer, pad_length), dtype=torch.long)

                    # If blank cell with no type info, then set to 0 tensor
                    if cell_type == 13:
                        y_tok[row_index, col_index, :] = torch.zeros(17, dtype=torch.long)

                    # Type info exists then
                    else:

                        # Store the metadata in the corresponding location in y_tok
                        y_tok[row_index, col_index, 0] = cell_type
                        y_tok[row_index, col_index, 1] = 0
                        y_tok[row_index, col_index, 2] = 0
                        y_tok[row_index, col_index, 3] = 2
                        y_tok[row_index, col_index, 4] = 0
                        y_tok[row_index, col_index, 5] = 12
                        y_tok[row_index, col_index, 6] = 0
                        y_tok[row_index, col_index, 7] = 0
                        y_tok[row_index, col_index, 8] = 0
                        y_tok[row_index, col_index, 9] = 0
                        y_tok[row_index, col_index, 10] = 0
                        y_tok[row_index, col_index, 11] = 0
                        y_tok[row_index, col_index, 12] = 0
                        y_tok[row_index, col_index, 13] = 0
                        y_tok[row_index, col_index, 14] = 0
                        y_tok[row_index, col_index, 15] = 0
                        y_tok[row_index, col_index, 16] = 0

    # Return the 3D LongTensor and the 3D LongTensor containing the metadata for each cell
    return x_tok, y_tok

def process_spreadsheet(file_path: str, tokenizer, max_rows=100, max_cols=100, pad_length=32) -> Optional[Tuple[torch.LongTensor, torch.LongTensor]]:
    """
    Processes a spreadsheet file (.xls, .xlsx, .csv) and extracts both the train tensor (x_tok) and metadata tensor (y_tok).

    Args:
        file_path (str): The path to the spreadsheet file.
        vocab: tokenizer object for encoding tokens.
        max_rows (int, optional): The maximum number of rows to process. Defaults to 100.
        max_cols (int, optional): The maximum number of columns to process. Defaults to 100.
        pad_length (int, optional): The length to which each tokenized cell value is padded. Defaults to 32.

    Returns:
        Optional[Tuple[torch.LongTensor, torch.LongTensor]]: A tuple containing:
            - A 100x100x32 PyTorch LongTensor with encoded tokenized values (x_tok).
            - A 100x100x15 PyTorch LongTensor containing the metadata for each cell (y_tok).
            Returns None if the file format is not supported or if an exception occurs during processing.
    """

    # Extract the file extension to determine the file type
    file_extension = file_path.split('.')[-1].lower()

    # Ensure the file is of a supported type (.xls, .xlsx, .csv)
    if file_extension not in ['xls', 'xlsx', 'csv']:
        return None  # Return None for unsupported file formats

    try:
        # Attempt to process the file based on its extension
        if file_extension == 'xls':
            # Call xls_metadata if the file is an .xls
            x_tok, y_tok = process_xls(file_path, tokenizer, max_rows, max_cols, pad_length)

        elif file_extension == 'xlsx':
            # Call xlsx_metadata if the file is an .xlsx
            x_tok, y_tok = process_xlsx(file_path, tokenizer, max_rows, max_cols, pad_length)

        elif file_extension == 'csv':
            # Call csv_metadata if the file is a .csv
            x_tok, y_tok = process_csv(file_path, tokenizer, max_rows, max_cols, pad_length)

    # If exception then just return None
    except Exception as e:
        return None
        # Attempt to process the file as a .csv if the initial processing fails
        # try:
        #     x_tok, y_tok = process_csv(file_path, tokenizer, max_rows, max_cols, pad_length)

        # except Exception:
        #     return None  # Return None if processing as .csv also fails

    # Return the tensors (x_tok and y_tok) if processing is successful
    return x_tok, y_tok

# def process_spreadsheet(file_path: str, tokenizer, max_rows=100, max_cols=100, pad_length=32) -> Optional[Tuple[torch.LongTensor, torch.LongTensor]]:
#     # Extract the file extension to determine the file type
#     file_extension = file_path.split('.')[-1].lower()

#     # Ensure the file is of a supported type (.xls, .xlsx, .csv)
#     if file_extension not in ['xls', 'xlsx', 'csv']:
#         return None  # Return None for unsupported file formats

#     try:
#         # Catch warnings within this block
#         with warnings.catch_warnings(record=True) as w:
#             warnings.simplefilter("always")
            
#             # Attempt to process the file based on its extension
#             if file_extension == 'xls':
#                 x_tok, y_tok = process_xls(file_path, tokenizer, max_rows, max_cols, pad_length)
#             elif file_extension == 'xlsx':
#                 x_tok, y_tok = process_xlsx(file_path, tokenizer, max_rows, max_cols, pad_length)
#             elif file_extension == 'csv':
#                 x_tok, y_tok = process_csv(file_path, tokenizer, max_rows, max_cols, pad_length)

#             # Check if the specific warning is in the list of captured warnings
#             for warning in w:
#                 if "Unknown extension is not supported and will be removed" in str(warning.message):
#                     print(f"Warning encountered in file: {file_path}")
#                     sys.exit(0)
#                     return None  # Stop further processing

#     except Exception as e:
#         print(f"Exception encountered in file {file_path}: {e}")
#         return None

#     # Return the tensors (x_tok and y_tok) if processing is successful
#     return x_tok, y_tok