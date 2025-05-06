# Import reload required funcs
import csv
import importlib
import itertools
import os
import warnings

import openpyxl
import pandas as pd
import torch
import xlrd
from IPython.display import display
from joblib import Parallel, delayed
from tqdm import tqdm
from xlrd.biffh import error_text_from_code
from xlrd.xldate import XLDateAmbiguous, XLDateError, xldate_as_tuple

# Custom imports
from utils import parseutil, selfutil

importlib.reload(parseutil)
importlib.reload(selfutil)
from utils.parseutil import (
    csv_dataType,
    h_cleanmetadata,
    h_tensors,
    xls_content,
    xls_mergedata,
    xls_metadata,
    xlsx_content,
    xlsx_dataType,
    xlsx_mergedata,
    xlsx_metadata,
)
from utils.selfutil import get_fileList

"""
@==========================================================================@
                              1. TOKENIZER 
@==========================================================================@
"""


# [1a] Tokenizer Data for XLS files
def xls_tokenize(file_path):

    # Open the workbook and get the first sheet
    workbook = xlrd.open_workbook(filename=file_path, formatting_info=True)

    # List to hold all cell contents
    sheet_contents = []

    df_dict = pd.read_excel(
        file_path,
        header=None,
        dtype=str,
        na_values=" ",
        keep_default_na=False,
        sheet_name=None,
        engine="xlrd",
    )

    # Loop over all sheets in the workbook
    for sheet_index, sheet in enumerate(workbook.sheets()):

        # Get the matching DataFrame for this sheet
        df_sheet = list(df_dict.values())[sheet_index]

        # Iterate over each cell using itertools.product
        for row_idx, col_idx in itertools.product(
            range(sheet.nrows), range(sheet.ncols)
        ):

            # Get the cell object
            cell = sheet.cell(row_idx, col_idx)

            # Get type and content using the helper function
            _, cell_content = xls_content(cell, workbook, df_sheet, row_idx, col_idx)

            # Append content to the contents list
            sheet_contents.append(cell_content)

    # Remove duplicates since tokenizer training
    sheet_contents = list(set(sheet_contents))

    # Return xls content
    return sheet_contents


# [1b] Tokenizer Data for XLSX files
def xlsx_tokenize(file_path):

    # Suppress all UserWarnings related to various xlsx issues
    warnings.filterwarnings("ignore", category=UserWarning)

    # Load the workbook with formulas resolved
    workbook = openpyxl.load_workbook(file_path, data_only=True)

    # Flat list to hold all cell contents across all sheets
    sheet_contents = []

    # Loop through all sheets
    for sheet in workbook.worksheets:

        # Iterate over all cells in the current sheet
        for row, col in itertools.product(
            range(sheet.max_row), range(sheet.max_column)
        ):

            # Adjust to 1-based indexing (openpyxl requirement)
            cell = sheet.cell(row=row + 1, column=col + 1)

            # Extract cell type
            cell_type = xlsx_dataType(cell.value, cell.number_format)

            # Use the cell type to get value
            cell_content = xlsx_content(cell_type, cell)

            # Append content to the flat list
            sheet_contents.append(cell_content)

    # Return deduplicate
    return list(set(sheet_contents))


# [1c] Tokenizer Data for CSV files
def csv_tokenize(file_path):

    # List to hold all cell contents
    sheet_contents = []

    # Get the delimiter from the file
    with open(file_path, mode="r", newline="", encoding="utf-8") as csvfile:
        delimiter = csv.Sniffer().sniff(csvfile.read()).delimiter

    # Go through the CSV file and read each row
    with open(file_path, mode="r", newline="", encoding="utf-8") as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            for cell in row:
                sheet_contents.append(cell)

    # Deduplicate the contents and return
    return list(set(sheet_contents))


def _process_single_file_for_tokenizer(file):
    try:
        # Get the file extension
        file_ext = file.split(".")[-1].lower()

        # Check the file type and call the appropriate function
        if file_ext == "xls":
            sheet_contents = xls_tokenize(file)
        elif file_ext == "xlsx":
            sheet_contents = xlsx_tokenize(file)
        elif file_ext == "csv":
            sheet_contents = csv_tokenize(file)
        else:
            print(f"Unsupported file type: {file_ext}")
            return []

        return sheet_contents

    except Exception as e:
        print(f"Error processing file {file}: {str(e)}")
        return []


# [1d] Final func to take dir and return all its sheets' data
def get_saffutok_data(dir_path, threads=1):

    # Get the list of valid/invalid files
    files, _ = get_fileList(dir_path)

    try:
        # Process files in parallel and get a list of lists
        results = Parallel(n_jobs=max(1, threads), timeout=99999)(
            delayed(_process_single_file_for_tokenizer)(f)
            for f in tqdm(files, desc="Tokenizing files", unit="file")
        )

        # Flatten all results directly and deduplicate in one go
        all_contents = set(
            content for file_contents in results for content in file_contents
        )

        # Print information
        print(f"Files/Tokens: {len(files)}/{len(all_contents)}")

        return list(all_contents)

    except Exception as e:
        raise e


"""
@==========================================================================@
                              1. DATALOADER 
@==========================================================================@
"""


# [2a] Helper function to setup the initial storage tensors
def init_tensors(max_rows=100, max_cols=100, pad_length=32, tokenizer=None):

    # Validate tokenizer
    if tokenizer is None:
        raise ValueError("Please provide tokenizer")

    # Initialize tensors
    y_tok = torch.zeros((max_rows, max_cols, 17), dtype=torch.long)
    x_tok = torch.zeros((max_rows, max_cols, pad_length), dtype=torch.long)
    x_masks = torch.zeros((max_rows, max_cols, pad_length), dtype=torch.long)

    # Setup default sequence (common for both modes)
    id_seq = torch.tensor(
        tokenizer.encode(
            [tokenizer._sod, tokenizer._eod] + [tokenizer._pad] * (pad_length - 2)
        ),
        dtype=torch.long,
    )
    mask_seq = torch.tensor([1, 1] + [0] * (pad_length - 2), dtype=torch.long)

    # Initialize tokens with default sequence
    x_tok = (
        id_seq.unsqueeze(0).unsqueeze(0).expand(max_rows, max_cols, pad_length).clone()
    )
    x_masks = (
        mask_seq.unsqueeze(0)
        .unsqueeze(0)
        .expand(max_rows, max_cols, pad_length)
        .clone()
    )

    return x_tok, x_masks, y_tok


# [2b] Helper function to take the dims and return tokpadded sentence
def saffu_pad_encode(sentence, row, col, x_tok, x_masks, pad_length=32, tokenizer=None):

    # Tokenize the sentence
    tokens = tokenizer.tokenize(sentence)

    # Truncate/pad and add start/end tokens and encode
    encoded_sentence = tokenizer.encode(
        [tokenizer._sod]
        + (
            tokens[: pad_length - 2]
            + [tokenizer._pad] * max(0, pad_length - 2 - len(tokens))
        )
        + [tokenizer._eod]
    )

    # Create the attention mask
    masks = [
        0 if token == tokenizer.encode([tokenizer._pad])[0] else 1
        for token in encoded_sentence
    ]

    # Set the tensors
    x_tok[row, col, :] = torch.tensor(encoded_sentence)
    x_masks[row, col, :] = torch.tensor(masks)

    # Return tensors
    return x_tok, x_masks


# [2b] DataLoader Processing for XLS files
def xls_dataloader(
    file_path, max_rows=100, max_cols=100, pad_length=32, tokenizer=None
):

    # Validate parameters and initialize tensors
    x_tok, x_masks, y_tok = init_tensors(max_rows, max_cols, pad_length, tokenizer)

    # Open the .xls file with formatting_info=True to get the metadata
    workbook = xlrd.open_workbook(filename=file_path, formatting_info=True)

    # Ensure BIFF version compatibility
    if workbook.biff_version != 80:
        raise ValueError(f"{workbook.biff_version}: {file_path}")

    # Access the first sheet in the workbook
    sheet = workbook.sheet_by_index(0)

    # Get the dataframe for parsing
    df_read = pd.read_excel(
        file_path,
        header=None,
        dtype=str,
        na_values=" ",
        keep_default_na=False,
        engine="xlrd",
    )

    # Resize dataframe to the required dimensions
    df = df_read.reindex(index=range(max_rows), columns=range(max_cols), fill_value="")

    # Iterate over each cell
    for row, col in itertools.product(
        range(min(sheet.nrows, max_rows)), range(min(sheet.ncols, max_cols))
    ):
        # Retrieve the cell object
        cell = sheet.cell(row, col)

        # Extract content using xls_content with correct arguments
        cell_type, cell_content = xls_content(cell, workbook, df, row, col)

        # Retrieve metadata using xls_metadata function
        cell_metadata = xls_metadata(cell, workbook, row, col, sheet, cell_type)

        # Encode cell value based on mode
        x_tok, x_masks = saffu_pad_encode(
            cell_content, row, col, x_tok, x_masks, pad_length, tokenizer
        )

        # Assign metadata dynamically using a loop
        for i, key in enumerate(cell_metadata):
            y_tok[row, col, i] = cell_metadata[key]

        # Clean up empty cells if needed
        x_tok, y_tok, x_masks = h_cleanmetadata(row, col, x_tok, y_tok, x_masks, True)

    # Propagate merged cell metadata
    x_tok, x_masks, y_tok = xls_mergedata(
        sheet, x_tok, y_tok, x_masks, max_rows, max_cols, True
    )

    # Return final tensors
    return x_tok, x_masks, y_tok


# [2c] DataLoader Processing for XLSX files
def xlsx_dataloader(
    file_path, max_rows=100, max_cols=100, pad_length=32, tokenizer=None
):

    # Suppress all UserWarnings related to various xlsx issues
    warnings.filterwarnings("ignore", category=UserWarning)

    # Validate parameters and get processing mode
    x_tok, x_masks, y_tok = init_tensors(max_rows, max_cols, pad_length, tokenizer)

    # Load the workbook and access the active sheet
    workbook = openpyxl.load_workbook(file_path, data_only=True)

    # Retrieve the sheet
    sheet = workbook.active

    # Iterate over each combination of rows and columns in the sheet, up to the 100th row and column
    for row, col in itertools.product(
        range(min(sheet.max_row, max_rows)), range(min(sheet.max_column, max_cols))
    ):
        # Adjust to 1-based indexing
        cell = sheet.cell(row=row + 1, column=col + 1)

        # Extract cell metadata
        cell_metadata = xlsx_metadata(cell, row, col, sheet)

        # Use the cell type to get value
        cell_content = xlsx_content(cell_metadata["type"], cell)

        # Encode cell value based on mode
        x_tok, x_masks = saffu_pad_encode(
            cell_content, row, col, x_tok, x_masks, pad_length, tokenizer
        )

        # Assign metadata dynamically using a loop
        for i, key in enumerate(cell_metadata):
            y_tok[row, col, i] = cell_metadata[key]

        # If it is an empty/blank cell then further processing is required
        x_tok, y_tok, x_masks = h_cleanmetadata(row, col, x_tok, y_tok, x_masks, True)

    # Handle merged cell ranges propogating values through em
    x_tok, y_tok, x_masks = xlsx_mergedata(
        sheet, x_tok, y_tok, x_masks, max_rows, max_cols, True
    )

    # Return the 3 tensors
    return x_tok, x_masks, y_tok


# [2d] DataLoader Processing for CSV files
def csv_dataloader(
    file_path, max_rows=100, max_cols=100, pad_length=32, tokenizer=None
):
    # Validate parameters and get processing mode
    x_tok, x_masks, y_tok = init_tensors(max_rows, max_cols, pad_length, tokenizer)

    # Get the delimiter from the file
    with open(file_path, mode="r", newline="", encoding="utf-8") as csvfile:
        delimiter = csv.Sniffer().sniff(csvfile.read()).delimiter

    # Open the file and read it as CSV using the detected delimiter
    with open(file_path, mode="r", newline="", encoding="utf-8") as file:
        reader = csv.reader(file, delimiter=delimiter)

        # Iterate through each row in the CSV file
        for row_index, row_data in enumerate(reader):
            if row_index >= max_rows:
                break

            # Iterate through each column in the row
            for col_index, cell_value in enumerate(row_data):
                if col_index >= max_cols:
                    break

                # Process cell if within limits
                if row_index < max_rows and col_index < max_cols:
                    # Get cell type and clean value
                    cell_val_str = cell_value.strip()
                    cell_type = csv_dataType(cell_val_str)

                    # Process based on mode
                    x_tok, x_masks = saffu_pad_encode(
                        cell_val_str,
                        row_index,
                        col_index,
                        x_tok,
                        x_masks,
                        pad_length,
                        tokenizer,
                    )

                    # Set metadata based on cell type
                    if cell_type == 13:  # Blank cell
                        y_tok[row_index, col_index, :] = torch.zeros(
                            17, dtype=torch.long
                        )
                    else:
                        # Standard CSV metadata setup
                        y_tok[row_index, col_index, 0] = cell_type
                        y_tok[row_index, col_index, 1] = 0  # No fill in CSV
                        y_tok[row_index, col_index, 2] = 0  # No alignment
                        y_tok[row_index, col_index, 3] = 2  # Default vertical align
                        y_tok[row_index, col_index, 4] = 0
                        y_tok[row_index, col_index, 5] = 11
                        y_tok[row_index, col_index, 6:17] = 0  # No formatting in CSV

    # Return finally
    return x_tok, x_masks, y_tok


# [2e] Final function to process the file and return tensors
def get_saffu_tensors(
    file_path, max_rows=100, max_cols=100, pad_length=32, tokenizer=None
):
    try:
        # Validate tokenizer
        if tokenizer is None:
            raise ValueError("Please provide tokenizer")

        # Get file extension and check if supported
        file_extension = file_path.split(".")[-1].lower()

        # Process based on file extension
        if file_extension == "xls":
            x_tok, x_masks, y_tok = xls_dataloader(
                file_path,
                max_rows=max_rows,
                max_cols=max_cols,
                pad_length=pad_length,
                tokenizer=tokenizer,
            )
        elif file_extension == "xlsx":
            x_tok, x_masks, y_tok = xlsx_dataloader(
                file_path,
                max_rows=max_rows,
                max_cols=max_cols,
                pad_length=pad_length,
                tokenizer=tokenizer,
            )
        elif file_extension == "csv":
            x_tok, x_masks, y_tok = csv_dataloader(
                file_path,
                max_rows=max_rows,
                max_cols=max_cols,
                pad_length=pad_length,
                tokenizer=tokenizer,
            )
        else:
            raise ValueError(
                f"Unsupported file format: {file_extension} for {file_path}"
            )

        return x_tok, x_masks, y_tok

    except Exception as e:
        raise e
