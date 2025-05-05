# Import self funcs
import importlib

from utils import selfutil

importlib.reload(selfutil)
import csv
import itertools
import re
import sys
import warnings
from datetime import datetime, time
from typing import Optional, Tuple

import openpyxl

# Import general usecase libraries
import pandas as pd
import torch

# Import parsing libraries
import xlrd
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from xlrd.xldate import XLDateAmbiguous, XLDateError, xldate_as_tuple

from utils.selfutil import tokenize_pad

"""
@==========================================================================@
                              1. COMMON
@==========================================================================@
"""


# [1a] Validates bert tokenizer or vocabulary then sets up tensor
def h_tensors(max_rows, max_cols, pad_length, tokenizer=None, vocab=None):
    """
    Initializes tensors for tokenized data, attention masks, and metadata.
    Supports BERT (tokenizer) or vocab-based processing.

    Args:
        max_rows (int): Maximum rows to process.
        max_cols (int): Maximum columns to process.
        pad_length (int): Length to pad/truncate sequences.
        tokenizer (optional): HuggingFace tokenizer for BERT-based processing.
        vocab (optional): Vocab object for non-BERT processing.

    Returns:
        tuple: (isBert, x_tok, x_masks, y_tok)
    """
    # Validate input parameters
    if tokenizer is None and vocab is None:
        raise ValueError("Either tokenizer or vocab must be provided")
    if tokenizer is not None and vocab is not None:
        raise ValueError("Only one of tokenizer or vocab should be provided")

    # Determine processing mode
    isBert = tokenizer is not None
    currobj = tokenizer if isBert else vocab

    # Initialize tensors
    y_tok = torch.zeros((max_rows, max_cols, 17), dtype=torch.long)
    x_tok = torch.zeros((max_rows, max_cols, pad_length), dtype=torch.long)
    x_masks = (
        torch.zeros((max_rows, max_cols, pad_length), dtype=torch.long)
        if isBert
        else torch.tensor([])
    )

    # Setup default sequence (common for both modes)
    minimal_ids = [currobj.cls_token_id, currobj.sep_token_id] + [
        currobj.pad_token_id
    ] * (pad_length - 2)
    minimal_mask = [1, 1] + [0] * (pad_length - 2)

    # Initialize tokens with default sequence
    for r in range(max_rows):
        for c in range(max_cols):
            # x_tok[r, c, :] = torch.tensor(minimal_ids, dtype=torch.long)

            if isBert:
                x_tok[r, c, :] = torch.tensor(minimal_ids, dtype=torch.long)
                x_masks[r, c, :] = torch.tensor(minimal_mask, dtype=torch.long)

    return isBert, x_tok, x_masks, y_tok


# [1b] Cleans up empty cell metadata and propogates value and metadata for centercontinuous cells
def h_cleanmetadata(row, col, x_tok, y_tok, x_masks, isBert):
    """
    Cleans metadata and tokenized data for empty/blank cells.

    Args:
        row (int): The row index of the cell.
        col (int): The column index of the cell.
        x_tok (torch.Tensor): Tensor storing tokenized values for each cell.
        y_tok (torch.Tensor): Tensor storing metadata for each cell.

    Returns:
        tuple: Updated (x_tok, y_tok) tensors after cleaning.
    """
    # If it is an empty/blank cell then further processing is required
    if y_tok[row, col, 0] in [0, 13, 14]:
        # Set font-related metadata to 0 (indices 4 to 10)
        y_tok[row, col, 4:11] = 0

        # If the cell is of the type center-across formatting
        if y_tok[row, col, 2] == 6:
            # Define variable to store col of where this center-across starts
            start_col = col

            # Loop until we find a cell that is centered across cells but not empty or blank
            while y_tok[row, start_col, 2] == 6 and y_tok[row, start_col, 0] in [0, 13]:
                # Move to the left
                start_col -= 1

            # Set the x_tok and y_tok tensor to these values
            x_tok[row, col, :] = x_tok[row, start_col, :].clone()
            y_tok[row, col, :] = y_tok[row, start_col, :].clone()
            if isBert:
                x_masks[row, col, :] = x_masks[row, start_col, :].clone()

    return x_tok, y_tok, x_masks


# [1c] Encodes cell value based on processing mode
def h_encode(cell_value, row, col, x_tok, x_masks, isBert, pad_length, currobj):
    """
    Encodes cell value using either BERT tokenizer or vocabulary and updates tensors.

    Args:
        cell_value (str): Value to encode
        row (int): Current row index
        col (int): Current column index
        x_tok (torch.Tensor): Token tensor to update
        x_masks (torch.Tensor): Mask tensor to update (used for BERT)
        isBert (bool): Whether using BERT mode
        pad_length (int): Length to pad sequences
        currobj: Either tokenizer or vocab object

    Returns:
        tuple: Updated (x_tok, x_masks)
    """
    if isBert:
        # Encode using BERT tokenizer
        encoded_value = currobj.encode_plus(
            cell_value,
            max_length=pad_length,
            truncation=True,
            padding="max_length",
            return_tensors="pt",
        )
        x_tok[row, col, :] = encoded_value["input_ids"][0]
        x_masks[row, col, :] = encoded_value["attention_mask"][0]
    else:
        # Encode using vocabulary
        x_tok[row, col, :] = torch.tensor(
            tokenize_pad(cell_value, currobj, pad_length), dtype=torch.long
        )

    return x_tok, x_masks


"""
@==========================================================================@
                               2. XLS
@==========================================================================@
"""


# [2a] Determines cell background color and fill status
def xls_fill(xf_record, workbook):
    """
    Determines whether the cell is filled with a background color.

    Args:
        xf_record (xlrd.formatting.XFRecord): Formatting information for the cell.
        workbook (xlrd.Book): Workbook object containing cell metadata.

    Returns:
        int: 1 if the cell is filled with a color other than white, 0 otherwise.
    """
    # Retrieve the color of the cell's background using the pattern_colour_index
    bg_color_rgb = workbook.colour_map.get(xf_record.background.pattern_colour_index)

    # Return 0 if the cell is not filled or filled with white; otherwise, return 1
    return 0 if bg_color_rgb in [None, (255, 255, 255)] else 1


# [2b] Gets merge status for cells in terms of v/h position
def xls_merge(row, col, sheet):
    """
    Determines the merge status of a cell in terms of its vertical and horizontal position.

    Args:
        row (int): The row index of the cell.
        col (int): The column index of the cell.
        sheet (xlrd.sheet.Sheet): The sheet object containing merged cells information.

    Returns:
        tuple: A tuple (horizontal_tag, vertical_tag) where:
            - horizontal_tag (int): 1 if at the beginning of horizontal range, 2 if inside, 0 otherwise.
            - vertical_tag (int): 1 if at the beginning of vertical range, 2 if inside, 0 otherwise.
    """
    # Initialize tags as 0 (outside) by default
    vertical_tag = 0
    horizontal_tag = 0

    # Iterate over the merged cells in the sheet
    for start_row, end_row, start_col, end_col in sheet.merged_cells:

        # Check if the current cell is within a merged range
        if start_row <= row < end_row and start_col <= col < end_col:

            # Determine vertical tag
            if row == start_row:
                vertical_tag = 1  # Beginning of vertical range
            else:
                vertical_tag = 2  # Inside vertical range

            # Determine horizontal tag
            if col == start_col:
                horizontal_tag = 1  # Beginning of horizontal range
            else:
                horizontal_tag = 2  # Inside horizontal range

            # Break the loop as the cell is part of a merged range
            break

    # Return the combined vertical and horizontal tags
    return horizontal_tag, vertical_tag


# [2c] Determines cell type and handles value parsing with formats
def xls_content(cell, workbook, df, row, col):
    """
    Determines the type of the cell, handles conversions, and retrieves the cell value.

    Args:
        cell (xlrd.sheet.Cell): Cell object to determine type and retrieve value.
        workbook (xlrd.Book): Workbook object containing cell metadata.
        df (pandas.DataFrame): DataFrame representation of the sheet for value extraction.
        row (int): The row index of the cell.
        col (int): The column index of the cell.

    Returns:
        tuple: A tuple (cell_type, cell_value) where:
            - cell_type (int): The type of the cell.
            - cell_value (str): The value of the cell.
    """
    xlrd_error_map = {
        0: "#NULL!",
        7: "#DIV/0!",
        15: "#VALUE!",
        23: "#REF!",
        29: "#NAME?",
        36: "#NUM!",
        42: "#N/A",
        43: "#GETTING_DATA",
    }

    # Empty cell with no formatting
    if cell.ctype == 0:
        return 0, ""

    # If text cell then check further to get exact text type
    elif cell.ctype == 1:
        format_str = workbook.format_map[
            workbook.xf_list[cell.xf_index].format_key
        ].format_str

        if "#,##0" in format_str:
            return 6, str(cell.value)  # Currency type

        if format_str == "0.00E+00":
            return 7, f"{cell.value}"  # Scientific subclass of text

        return 1, str(cell.value)  # Text type

    # If numeric cell then check further to get exact numeric type
    elif cell.ctype == 2:
        format_str = workbook.format_map[
            workbook.xf_list[cell.xf_index].format_key
        ].format_str

        if format_str == "0.00E+00":
            return 7, f"{cell.value:.2e}"  # Scientific subclass of number

        elif "#,##0" in format_str:

            if "$€" in format_str:
                return 6, "€" + str(float(cell.value))

            return 6, "$" + str(float(cell.value))

        elif "%" in format_str:
            return 5, str(cell.value * 100) + "%"  # Percentage type

        elif isinstance(cell.value, float):
            if cell.value.is_integer():
                return 3, str(df.iat[row, col])  # Integer type
            return 4, str(df.iat[row, col])  # Float type

        return 2, str(df.iat[row, col])  # General numeric type

    # If date cell then check further to get exact date type
    elif cell.ctype == 3:
        try:
            date_tuple = xldate_as_tuple(cell.value, workbook.datemode)

            if date_tuple[:3] == (0, 0, 0):
                return 9, str(df.iat[row, col])  # Time type

            elif date_tuple[3:] == (0, 0, 0):
                return 8, str(df.iat[row, col]).split(" ")[0]  # Date type

            return 10, str(df.iat[row, col]).replace(
                "1900-", "2000-", 1
            )  # Datetime type

        except (XLDateError, XLDateAmbiguous):
            return 10, str(df.iat[row, col]).replace(
                "1900-", "2000-", 1
            )  # Datetime type

    # Boolean cell
    elif cell.ctype == 4:
        return 11, str(df.iat[row, col])

    # Error cell
    elif cell.ctype == 5:
        return 12, xlrd_error_map.get(cell.value, f"#ERR{cell.value}")

    # Blank cell (Empty cell with formatting)
    elif cell.ctype == 6:
        return 13, ""

    # Unknown type not in our keys return by default
    return 14, ""


# [2d] Handles propogation of metadata and data after loop is done for remaining merged cells
def xls_mergedata(sheet, x_tok, y_tok, x_masks, max_rows, max_cols, isBert):
    """
    Handles processing and propagating data across merged cells.

    Args:
        sheet (xlrd.sheet.Sheet): The sheet object containing merged cells information.
        x_tok (torch.Tensor): Tensor storing tokenized values for each cell.
        y_tok (torch.Tensor): Tensor storing metadata for each cell.
        max_rows (int): Maximum number of rows to process.
        max_cols (int): Maximum number of columns to process.

    Returns:
        tuple: Updated (x_tok, y_tok) tensors after handling merged cells.
    """
    # Iterate over the merged cells in the sheet
    for start_row, end_row, start_col, end_col in sheet.merged_cells:
        # Derive adjusted indices for accessing the combined cell
        adj_sr = min(start_row, max_rows - 1)
        adj_er = min(end_row - 1, max_rows - 1)  # End row index is inclusive
        adj_sc = min(start_col, max_cols - 1)
        adj_ec = min(end_col - 1, max_cols - 1)  # End column index is inclusive

        # Extract the starting values from the top-left corner of the merged range
        x_tok_start = x_tok[adj_sr, adj_sc, :].clone()
        y_tok_start = y_tok[adj_sr, adj_sc, :].clone()

        if isBert:
            x_masks_start = x_masks[adj_sr, adj_sc, :].clone()

        # Propagate values across the merged range
        for row, col in itertools.product(
            range(adj_sr, adj_er + 1), range(adj_sc, adj_ec + 1)
        ):
            x_tok[row, col, :] = x_tok_start

            if isBert:
                x_masks[row, col, :] = x_masks_start

            # Duplicate y_tok values excluding indices 15 and 16
            y_tok[row, col, :15] = y_tok_start[:15]
            y_tok[row, col, 17:] = y_tok_start[17:]

    return x_tok, x_masks, y_tok


# [2e] Function to retrieve xls cell's metadata
def xls_metadata(cell, workbook, row, col, sheet, cell_type):
    """
    Extracts metadata from an XLS cell and returns it as a dictionary.

    Args:
        cell (xlrd.sheet.Cell): The cell object from the XLS sheet.
        workbook (xlrd.Book): The XLS workbook object.
        row (int): Row index of the cell.
        col (int): Column index of the cell.
        sheet (xlrd.sheet.Sheet): The XLS sheet object.

    Returns:
        dict: A dictionary (`cell_metadata`) containing 17 metadata fields for the cell.
    """
    # Access the XFRecord object using the cell's XF index
    xf_record = workbook.xf_list[cell.xf_index]

    # Retrieve color information
    rgb = workbook.colour_map.get(workbook.font_list[xf_record.font_index].colour_index)

    # Construct metadata dictionary in the same order as manual y_tok assignments
    cell_metadata = {
        "type": cell_type,  # Type of cell
        "fill": xls_fill(xf_record, workbook),  # Background fill
        "halign": xf_record.alignment.hor_align,  # Horizontal alignment
        "valign": xf_record.alignment.vert_align,  # Vertical alignment
        "font_family": workbook.font_list[xf_record.font_index].family,  # Font family
        "font_size": workbook.font_list[xf_record.font_index].height // 20,  # Font size
        "bold": workbook.font_list[xf_record.font_index].bold,  # Bold flag
        "italic": workbook.font_list[xf_record.font_index].italic,  # Italic flag
        "underline": workbook.font_list[
            xf_record.font_index
        ].underline_type,  # Underline
        "esc": workbook.font_list[
            xf_record.font_index
        ].escapement,  # Superscript/subscript
        "font_color": (
            1 if (rgb and rgb != (0, 0, 0)) else 0
        ),  # Text color (1 if colored, else 0)
        "btop": xf_record.border.top_line_style,  # Top border style
        "bbottom": xf_record.border.bottom_line_style,  # Bottom border style
        "bleft": xf_record.border.left_line_style,  # Left border style
        "bright": xf_record.border.right_line_style,  # Right border style
        "hmerge": xls_merge(row, col, sheet)[0],  # Horizontal merge
        "vmerge": xls_merge(row, col, sheet)[1],  # Vertical merge
    }

    return cell_metadata


# [2f] Main func for processing xls files
def process_xls(
    file_path, max_rows=100, max_cols=100, pad_length=32, *, tokenizer=None, vocab=None
):
    """
    Extracts metadata for each cell in the first sheet of the given spreadsheet and tokenizes cell values.

    Args:
        file_path (str): Path to the spreadsheet file.
        vocab (dict): Vocabulary mapping tokens to indices.
        max_rows (int): Maximum rows to process.
        max_cols (int): Maximum columns to process.
        pad_length (int): Length to pad tokenized sequences.

    Returns:
        tuple: A tuple (x_tok, y_tok) where:
            - x_tok (torch.Tensor): A 3D tensor of tokenized cell values.
            - y_tok (torch.Tensor): A 3D tensor of metadata for each cell.
    """

    # Validate parameters and initialize tensors
    isBert, x_tok, x_masks, y_tok = h_tensors(
        max_rows, max_cols, pad_length, tokenizer, vocab
    )

    # Open the .xls file with formatting_info=True to get the metadata
    workbook = xlrd.open_workbook(filename=file_path, formatting_info=True)

    # Ensure BIFF version compatibility
    if workbook.biff_version != 80:
        raise ValueError(f"{workbook.biff_version}: {file_path}")

    # Access the first sheet in the workbook
    sheet = workbook.sheet_by_index(0)

    # Get the dataframe for parsing
    df_read = pd.read_excel(
        file_path,
        header=None,
        dtype=str,
        na_values=" ",
        keep_default_na=False,
        engine="xlrd",
    )

    # Resize dataframe to the required dimensions
    df = df_read.reindex(index=range(max_rows), columns=range(max_cols), fill_value="")

    # Iterate over each cell
    for row, col in itertools.product(
        range(min(sheet.nrows, max_rows)), range(min(sheet.ncols, max_cols))
    ):
        # Retrieve the cell object
        cell = sheet.cell(row, col)

        # Extract content using xls_content with correct arguments
        cell_type, cell_content = xls_content(cell, workbook, df, row, col)

        # Retrieve metadata using xls_metadata function
        cell_metadata = xls_metadata(cell, workbook, row, col, sheet, cell_type)

        # Encode cell value based on mode
        x_tok, x_masks = h_encode(
            cell_content,
            row,
            col,
            x_tok,
            x_masks,
            isBert,
            pad_length,
            tokenizer if isBert else vocab,
        )

        # Assign metadata dynamically using a loop
        for i, key in enumerate(cell_metadata):
            y_tok[row, col, i] = cell_metadata[key]

        # Clean up empty cells if needed
        x_tok, y_tok, x_masks = h_cleanmetadata(row, col, x_tok, y_tok, x_masks, isBert)

    # Propagate merged cell metadata
    x_tok, x_masks, y_tok = xls_mergedata(
        sheet, x_tok, y_tok, x_masks, max_rows, max_cols, isBert
    )

    # Return final tensors
    return x_tok, x_masks, y_tok


"""
@==========================================================================@
                               3. XLSX
@==========================================================================@
"""


# [3a] Get the type of data
def xlsx_dataType(value, number_format):
    """
    Determines the type of the cell and handles necessary conversions.

    Int to Type Key Mapping:
        0  -> Empty cell
        1  -> Text cell
        2  -> Numeric type
        3  -> Integer subclass of number
        4  -> Float subclass of number
        5  -> Percentage subclass of number
        6  -> Currency subclass of number
        7  -> Scientific subclass of number
        8  -> Date type
        9  -> Time subclass of date
        10 -> Datetime subclass of date
        11 -> Boolean type with T or F value
        12 -> Error type corresponding to #REF!, #VALUE! etc. in excel
        13 -> Blank cell with formatting
        14 -> Unknown type not in our keys

    Args:
        value: The value of the cell.
        number_format (str): The number format of the cell.

    Returns:
        int: The determined type of the cell (0-14).
    """
    # Empty cell with no value return blank for this
    if value is None:
        return 13

    # Boolean cell
    elif isinstance(value, bool):
        return 11

    # Error cell based on certain error values in Excel
    elif any(
        error in str(value)
        for error in {
            "#VALUE!",
            "#REF!",
            "#DIV/0!",
            "#NAME?",
            "#N/A",
            "#NUM!",
            "#NULL!",
        }
    ):
        return 12

    # Currency format cell
    elif "#,##0" in number_format:
        return 6

    # Percentage format cell
    elif number_format.endswith("%"):
        return 5

    # Check for scientific notation format
    elif "E+" in number_format or "E-" in number_format:
        return 7

    # If float type
    elif isinstance(value, float):
        # Check if it's an integer subclass
        if value.is_integer():
            return 3
        # Otherwise, it's a float
        return 4

    # If integer type
    elif isinstance(value, int):
        return 3

    # Time type cell
    elif isinstance(value, time):
        return 9

    # Datetime type cell
    elif isinstance(value, datetime):
        # Check if the time part is exactly midnight
        if value.time() == time(0, 0):
            # If the datetime is formatted as a date, return it as a date type
            if "d" in number_format.lower() and "h" not in number_format.lower():
                return 8
            else:
                # If there is significant time information, return it as datetime
                return 10
        else:
            # If the time part is not midnight, it's definitely a datetime
            return 10

    # Blank cell with formatting
    elif value == "":
        return 13

    # Text cell
    elif isinstance(value, str):
        return 1

    # Return 14 for any unknown type
    return 14


# [3b] Get the parsed value of data in cell
def xlsx_content(cell_type, cell):
    # If empty cell then return an empty string
    if cell_type == 0:
        return ""

    # Text cell
    elif cell_type == 1:
        return str(cell.value)

    # Numeric cell default type
    elif cell_type == 2:
        return str(cell.value)

    # Integer subclass of number, remove proceeding 0s and return str
    elif cell_type == 3:
        return str(cell.value)

    # Decimal subclass then convert to float and return str
    elif cell_type == 4:
        return str(cell.value)

    # Percentage subclass then convert to value add % sign and return the string
    elif cell_type == 5:
        return str(cell.value * 100) + "%"

    # Currency subclass then convert to value add currency symbol and return the string
    elif cell_type == 6:
        # Check if cell value already is a string so it will have the symbol
        if isinstance(cell.value, str):
            return str(cell.value)
        # Check if euro sign in format string because that is separate
        if "$€" in cell.number_format:
            return "€" + str(float(cell.value))
        # In other case for usd values we gotta add the $ symbol
        return "$" + str(float(cell.value))

    # If scientific subclass then convert to value add 'E' sign and return the string
    elif cell_type == 7:
        if isinstance(cell.value, (int, float)):
            return f"{cell.value:.2e}"
        else:
            return str(cell.value)

    # If date subclass then use the df to return value
    elif cell_type == 8:
        return str(cell.value).split(" ")[0]

    elif cell_type in [9, 10]:
        return str(cell.value)

    # If boolean cell then use df to get string
    elif cell_type == 11:
        return str(cell.value)

    # If error cell then use df to get string
    elif cell_type == 12:
        return str(cell.value)

    # If blank cell (Empty cell with formatting) then return an empty string
    elif cell_type == 13:
        return ""

    # Return empty string for default case
    return ""


# [3c] Get the alignment standardized keys for cell
def xlsx_alignment(cell):
    # Define dict to store the alignment keys mapped to xlrd
    openpyxl_to_xlrd = {
        "horizontal": {
            "general": 0,
            "left": 1,
            "center": 2,
            "right": 3,
            "fill": 4,
            "justify": 5,
            "centerContinuous": 6,
            "distributed": 7,
        },
        "vertical": {
            "top": 0,
            "center": 1,
            "bottom": 2,
            "justify": 3,
            "distributed": 4,
        },
    }

    # Get the horizontal and vertical alignment else default if key not found
    horiz = openpyxl_to_xlrd["horizontal"].get(cell.alignment.horizontal, 0)
    vert = openpyxl_to_xlrd["vertical"].get(cell.alignment.vertical, 2)

    # Concatenate the keys and return as int
    return horiz, vert


# [3d] Get the borders standardized keys of cell
def xlsx_borders(cell):
    # Define dict to store the border keys mapped to xlrd
    openpyxl_to_xlrd = {
        "none": 0,
        "thin": 1,
        "medium": 2,
        "dashed": 3,
        "dotted": 4,
        "thick": 5,
        "double": 6,
        "hair": 7,
        "mediumDashed": 8,
        "dashDot": 9,
        "mediumDashDot": 10,
        "dashDotDot": 11,
        "mediumDashDotDot": 12,
        "slantDashDot": 13,
    }

    return (
        openpyxl_to_xlrd.get(cell.border.top.style or "none"),
        openpyxl_to_xlrd.get(cell.border.bottom.style or "none"),
        openpyxl_to_xlrd.get(cell.border.left.style or "none"),
        openpyxl_to_xlrd.get(cell.border.right.style or "none"),
    )


# [3e] Get the merge tags for cell
def xlsx_merge(row, col, sheet):
    """
    Determines the merge status of a cell in terms of its vertical and horizontal position within a merged range.

    Int to Merge Status Mapping:
        1 -> Outside any merged range ('O')
        2 -> Beginning of a merged range ('B')
        3 -> Inside a merged range but not at the beginning ('I')

    Args:
        row (int): The row index of the cell.
        col (int): The column index of the cell.
        sheet (openpyxl.worksheet.worksheet.Worksheet): The sheet object containing the merged cells information.

    Returns:
        int: A 2-digit integer where the first digit represents the vertical position
             and the second digit represents the horizontal position within a merged range.
    """
    # Initialize tags as 1 (outside) by default
    vertical_tag = 0
    horizontal_tag = 0

    # Convert 0-based indexing to 1-based indexing for openpyxl
    row_1_based = row + 1
    col_1_based = col + 1
    cell_coord = openpyxl.utils.get_column_letter(col_1_based) + str(row_1_based)

    # Iterate over the merged cells in the sheet
    for merged in sheet.merged_cells.ranges:
        # Check if the current cell is within a merged range
        if cell_coord in merged:
            # Determine vertical tag
            if row_1_based == merged.min_row:
                vertical_tag = 1  # Beginning of vertical range
            else:
                vertical_tag = 2  # Inside vertical range

            # Determine horizontal tag
            if col_1_based == merged.min_col:
                horizontal_tag = 1  # Beginning of horizontal range
            else:
                horizontal_tag = 2  # Inside horizontal range

            # Break the loop as the cell is part of a merged range
            break

    # Return the combined vertical and horizontal tags as a 2-digit integer
    return horizontal_tag, vertical_tag


# [3f] Get the fontcolor of cell
def xlsx_fontcol(cell):
    # Check if None color
    if cell.font.color is None:
        return 0

    # Check tint first of all
    if float(cell.font.color.tint) != 0.0:
        return 1

    # Check theme
    if cell.font.color.theme is not None and cell.font.color.theme != 1:
        return 1

    # Check if rgb exists and is not None
    if cell.font.color.rgb is not None:
        # Ensure the RGB value is 8 characters long (including opacity) and starts with 'FF' for full opacity
        if isinstance(cell.font.color.rgb, str) and len(cell.font.color.rgb) == 8:
            # Extract the last 6 characters which represent the color and check if it's not black
            font_color = f"#{cell.font.color.rgb[ 2: ].lower()}"

            # If color is not black return 1 else 0
            if font_color != "#000000":
                return 1  # Font is colored
            else:
                return 0  # Font is black (default)
        else:
            return 0  # RGB value doesn't match expected format

    # If rgb is None
    else:
        return 0


# [3g] Propogate data through merged cells
def xlsx_mergedata(sheet, x_tok, y_tok, x_masks, max_rows, max_cols, isBert):
    """
    Handles processing and propagating data across merged cells, including attention masks if in BERT mode.

    Args:
        sheet: The active worksheet being processed.
        x_tok (torch.Tensor): Tensor storing tokenized values for each cell.
        y_tok (torch.Tensor): Tensor storing metadata for each cell.
        x_masks (torch.Tensor): Tensor storing attention masks for each cell (optional for BERT mode).
        max_rows (int): Maximum number of rows to process.
        max_cols (int): Maximum number of columns to process.
        isBert (bool): Flag indicating if processing is in BERT mode.

    Returns:
        tuple: Updated (x_tok, y_tok, x_masks) tensors after handling merged cells.
    """
    # Loop through merged cell ranges
    for merged in sheet.merged_cells.ranges:
        # Parse the range into start and end coordinates
        start_cell, end_cell = str(merged).split(":")
        start_col, start_row = coordinate_from_string(start_cell)
        end_col, end_row = coordinate_from_string(end_cell)

        # Convert to 0-based indexing
        start_row = max(0, start_row - 1)
        start_col = max(0, column_index_from_string(start_col) - 1)
        end_row = min(end_row - 1, max_rows - 1)
        end_col = min(column_index_from_string(end_col) - 1, max_cols - 1)

        # Skip invalid ranges
        if start_row >= max_rows or start_col >= max_cols:
            continue

        # Extract the starting values
        x_tok_start = x_tok[start_row, start_col, :].clone()
        y_tok_start = y_tok[start_row, start_col, :].clone()
        if isBert:
            x_masks_start = x_masks[start_row, start_col, :].clone()

        # Propagate values across the merged range
        for rr, cc in itertools.product(
            range(start_row, end_row + 1), range(start_col, end_col + 1)
        ):
            if rr < max_rows and cc < max_cols:  # Clip bounds
                x_tok[rr, cc, :] = x_tok_start
                y_tok[rr, cc, :15] = y_tok_start[:15]
                y_tok[rr, cc, 17:] = y_tok_start[17:]
                if isBert:
                    x_masks[rr, cc, :] = x_masks_start

    return x_tok, y_tok, x_masks


# [3h] Function to retrieve xlsx cell's metadata
def xlsx_metadata(cell, row, col, sheet):
    """
    Extracts metadata from an XLSX cell and returns it as a dictionary.

    Args:
        cell: The cell object from the XLSX sheet.
        row (int): Row index of the cell.
        col (int): Column index of the cell.
        sheet: The XLSX sheet object.

    Returns:
        dict: A dictionary (`cell_metadata`) containing 17 metadata fields for the cell.
    """
    # Get alignments/borders/merged ranges since the funcs give em together
    halign, valign = xlsx_alignment(cell)
    btop, bbot, bleft, bright = xlsx_borders(cell)
    hmerge, vmerge = xlsx_merge(row, col, sheet)

    # Construct metadata dictionary in the same order as manual y_tok assignments
    cell_metadata = {
        "type": xlsx_dataType(cell.value, cell.number_format),
        "fill": (
            0
            if (
                cell.fill.start_color.index
                if cell.fill and cell.fill.start_color
                else None
            )
            in ["00000000", "00FFFFFF", 0]
            else 1
        ),
        "halign": halign,
        "valign": valign,
        "font_family": (int(cell.font.family) if cell.font.family else 0),
        "font_size": cell.font.size,
        "bold": int(cell.font.bold),
        "italic": int(cell.font.italic),
        "underline": (
            1
            if cell.font.underline == "single"
            else 2 if cell.font.underline == "double" else 0
        ),
        "esc": (
            1
            if cell.font.vertAlign == "superscript"
            else 2 if cell.font.vertAlign == "subscript" else 0
        ),
        "font_color": xlsx_fontcol(cell),
        "btop": btop,
        "bbottom": bbot,
        "bleft": bleft,
        "bright": bright,
        "hmerge": hmerge,
        "vmerge": vmerge,
    }

    return cell_metadata


# [3i] Main processing func
def process_xlsx(
    file_path, max_rows=100, max_cols=100, pad_length=32, *, tokenizer=None, vocab=None
):

    # Suppress all UserWarnings related to various xlsx issues
    warnings.filterwarnings("ignore", category=UserWarning)

    # Validate parameters and get processing mode
    isBert, x_tok, x_masks, y_tok = h_tensors(
        max_rows, max_cols, pad_length, tokenizer, vocab
    )

    # Load the workbook and access the active sheet
    workbook = openpyxl.load_workbook(file_path, data_only=True)

    # Retrieve the sheet
    sheet = workbook.active

    # Iterate over each combination of rows and columns in the sheet, up to the 100th row and column
    for row, col in itertools.product(
        range(min(sheet.max_row, max_rows)), range(min(sheet.max_column, max_cols))
    ):
        # Adjust to 1-based indexing
        cell = sheet.cell(row=row + 1, column=col + 1)

        # Extract cell metadata
        cell_metadata = xlsx_metadata(cell, row, col, sheet)

        # Use the cell type to get value
        cell_content = xlsx_content(cell_metadata["type"], cell)

        # Encode cell value based on mode
        x_tok, x_masks = h_encode(
            cell_content,
            row,
            col,
            x_tok,
            x_masks,
            isBert,
            pad_length,
            tokenizer if isBert else vocab,
        )

        # Assign metadata dynamically using a loop
        for i, key in enumerate(cell_metadata):
            y_tok[row, col, i] = cell_metadata[key]

        # If it is an empty/blank cell then further processing is required
        x_tok, y_tok, x_masks = h_cleanmetadata(row, col, x_tok, y_tok, x_masks, isBert)

    # Handle merged cell ranges propogating values through em
    x_tok, y_tok, x_masks = xlsx_mergedata(
        sheet, x_tok, y_tok, x_masks, max_rows, max_cols, isBert
    )

    # Return the 2D NumPy array containing the metadata for each cell
    return x_tok, x_masks, y_tok


"""
@==========================================================================@
                               4. CSV
@==========================================================================@
"""


# [4a] Gets the type of data in cell
def csv_dataType(value: str) -> int:
    """
    Determines the type of the cell and handles necessary conversions.

    Int to Type Key Mapping:
        0  -> Empty cell
        1  -> Text cell
        2  -> Numeric type
        3  -> Integer subclass of number
        4  -> Float subclass of number
        5  -> Percentage subclass of number
        6  -> Currency subclass of number
        7  -> Scientific subclass of number
        8  -> Date type
        9  -> Time subclass of date
        10 -> Datetime subclass of date
        11 -> Boolean type with T or F value
        12 -> Error type corresponding to #REF!, #VALUE! etc. in Excel
        13 -> Blank cell with formatting
        14 -> Unknown type not in our keys

    Args:
        value (str): The value of the cell.

    Returns:
        int: The determined type of the cell (0-14).
    """

    # Check for blank cell
    if value == "":
        return 13

    # Check for date type
    date_patterns = [
        r"^\d{1,2}/\d{1,2}/\d{2,4}$",
        r"^\d{1,2}-\w{3}-\d{2,4}$",
        r"^\w{3,9}-\d{2}$",
        r"^\w{3,9}\s\d{1,2},\s\d{4}$",
        r"^\d{4}-\d{1,2}-\d{1,2}$",
    ]
    time_patterns = [
        r"^\d{1,2}:\d{2}(:\d{2})?(\.\d+)?\s?(AM|PM)?$",
        r"^\d{1,2}:\d{2}\s?(AM|PM)?$",
    ]
    datetime_patterns = [
        r"^\d{1,2}/\d{1,2}/\d{2,4}\s\d{1,2}:\d{2}(:\d{2})?\s?(AM|PM)?$",  # 1/20/00 12:33 AM, 1/20/00 12:33:00
        r"^\d{1,2}-\w{3}-\d{2,4}\s\d{1,2}:\d{2}(:\d{2})?\s?(AM|PM)?$",  # 1-Jan-00 12:33 AM
        r"^\d{1,2}/\d{1,2}/\d{4}\s\d{1,2}:\d{2}(:\d{2})?\s?(AM|PM)?$",  # 1/20/2000 12:33:00
    ]

    # Check for datetime first
    if any(re.fullmatch(dp, value) for dp in datetime_patterns):
        return 10

    # Check for date
    if any(re.fullmatch(dp, value.split()[0]) for dp in date_patterns):
        # Check for datetime
        if any(re.fullmatch(tp, value.split()[-1]) for tp in time_patterns):
            return 10
        else:
            return 8

    # Check for time
    if any(re.fullmatch(tp, value) for tp in time_patterns):
        return 9

    # Check for percentage
    if value.endswith("%") and re.match(r"^-?\d+(\.\d+)?%$", value):
        return 5

    # Check for currency
    if re.match(r"^[-\(\{\[]?\s*[\$£€₹¥]?\s*\d+(\.\d{2})?\s*[\)\}\]]?$", value):
        return 6

    # Check for scientific notation
    if re.match(r"^[+-]?\d+(\.\d+)?[eE][+-]?\d+$", value):
        return 7

    # Check for float type
    if re.match(r"^[+-]?\d+\.\d+$", value):
        return 4

    # Check for integer type
    if re.match(r"^[+-]?\d+$", value):
        return 3

    # Check for boolean
    if value.lower() in {"true", "false"}:
        return 11

    # Check for error type
    if value in {"#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}:
        return 12

    # Check for str
    if isinstance(value, str):
        return 1

    # Default to unk if none of the above
    return 14


# [4c] Main processing function
def process_csv(
    file_path, max_rows=100, max_cols=100, pad_length=32, *, tokenizer=None, vocab=None
):
    """
    Extracts metadata and tokenized values for each cell in a CSV file.
    Supports both BERT tokenizer and vocabulary-based processing.

    Args:
        file_path (str): Path to the CSV file.
        max_rows (int, optional): Maximum rows to process. Defaults to 100.
        max_cols (int, optional): Maximum columns to process. Defaults to 100.
        pad_length (int, optional): Length to pad sequences. Defaults to 32.
        tokenizer: BERT tokenizer object (optional)
        vocab: Vocabulary object (optional)

    Returns:
        tuple: (x_tok, x_masks, y_tok) where:
            - x_tok: Tensor of tokenized values
            - x_masks: Attention mask tensor (empty if using vocab)
            - y_tok: Metadata tensor
    """
    # Validate parameters and get processing mode
    isBert, x_tok, x_masks, y_tok = h_tensors(
        max_rows, max_cols, pad_length, tokenizer, vocab
    )

    # Get the delimiter from the file
    with open(file_path, mode="r", newline="", encoding="utf-8") as csvfile:
        delimiter = csv.Sniffer().sniff(csvfile.read()).delimiter

    # Open the file and read it as CSV using the detected delimiter
    with open(file_path, mode="r", newline="", encoding="utf-8") as file:
        reader = csv.reader(file, delimiter=delimiter)

        # Iterate through each row in the CSV file
        for row_index, row_data in enumerate(reader):
            if row_index >= max_rows:
                break

            # Iterate through each column in the row
            for col_index, cell_value in enumerate(row_data):
                if col_index >= max_cols:
                    break

                # Process cell if within limits
                if row_index < max_rows and col_index < max_cols:
                    # Get cell type and clean value
                    cell_val_str = cell_value.strip()
                    cell_type = csv_dataType(cell_val_str)

                    # Process based on mode
                    x_tok, x_masks = h_encode(
                        cell_val_str,
                        row_index,
                        col_index,
                        x_tok,
                        x_masks,
                        isBert,
                        pad_length,
                        tokenizer if isBert else vocab,
                    )

                    # Set metadata based on cell type
                    if cell_type == 13:  # Blank cell
                        y_tok[row_index, col_index, :] = torch.zeros(
                            17, dtype=torch.long
                        )
                    else:
                        # Standard CSV metadata setup
                        y_tok[row_index, col_index, 0] = cell_type
                        y_tok[row_index, col_index, 1] = 0  # No fill in CSV
                        y_tok[row_index, col_index, 2] = 0  # No alignment
                        y_tok[row_index, col_index, 3] = 2  # Default vertical align
                        y_tok[row_index, col_index, 4] = 0
                        y_tok[row_index, col_index, 5] = 11
                        y_tok[row_index, col_index, 6:17] = 0  # No formatting in CSV

    # Return finally
    return x_tok, x_masks, y_tok


"""
@==========================================================================@
                               5. Main Processor
@==========================================================================@
"""


# [5a] Main processor func to run file by file
def process_spreadsheet(
    file_path: str,
    max_rows=100,
    max_cols=100,
    pad_length=32,
    *,
    tokenizer=None,
    vocab=None,
):
    """
    Processes spreadsheet files (.xls, .xlsx, .csv) with BERT tokenizer or vocabulary-based processing.

    Args:
        file_path (str): Path to the spreadsheet file to process
        max_rows (int, optional): Maximum number of rows to process. Defaults to 100.
        max_cols (int, optional): Maximum number of columns to process. Defaults to 100.
        pad_length (int, optional): Length to pad/truncate sequences to. Defaults to 32.
        tokenizer: BERT tokenizer object for BERT-based processing (optional)
        vocab: Vocabulary object for vocabulary-based processing (optional)

    Returns:
        tuple: A tuple containing:
            - x_tok (torch.Tensor): Token IDs tensor of shape (max_rows, max_cols, pad_length)
            - x_masks (torch.Tensor): Attention masks tensor of shape (max_rows, max_cols, pad_length)
                                    if using BERT, empty tensor if using vocabulary
            - y_tok (torch.Tensor): Metadata tensor of shape (max_rows, max_cols, 17)

    Raises:
        ValueError: If neither or both tokenizer/vocab provided, or if file format unsupported
        Exception: For other processing errors during spreadsheet handling

    Notes:
        - Must provide exactly one of tokenizer or vocab
        - Supported file formats: .xls, .xlsx, .csv
        - Processing mode determined by which tokenization object provided
        - Returns 3 tensors regardless of mode (x_masks empty for vocab mode)
    """
    try:
        # Validate input parameters
        if tokenizer is None and vocab is None:
            raise ValueError("Either tokenizer or vocab must be provided")
        if tokenizer is not None and vocab is not None:
            raise ValueError("Only one of tokenizer or vocab should be provided")

        # Determine processing mode
        isBert = tokenizer is not None

        # Get file extension and check if supported
        file_extension = file_path.split(".")[-1].lower()

        # Process based on file extension
        if file_extension == "xls":
            x_tok, x_masks, y_tok = process_xls(
                file_path,
                max_rows=max_rows,
                max_cols=max_cols,
                pad_length=pad_length,
                tokenizer=tokenizer,
                vocab=vocab,
            )
        elif file_extension == "xlsx":
            x_tok, x_masks, y_tok = process_xlsx(
                file_path,
                max_rows=max_rows,
                max_cols=max_cols,
                pad_length=pad_length,
                tokenizer=tokenizer,
                vocab=vocab,
            )
        elif file_extension == "csv":
            x_tok, x_masks, y_tok = process_csv(
                file_path,
                max_rows=max_rows,
                max_cols=max_cols,
                pad_length=pad_length,
                tokenizer=tokenizer,
                vocab=vocab,
            )
        else:
            raise ValueError(
                f"Unsupported file format: {file_extension} for {file_path}"
            )

        return x_tok, x_masks, y_tok

    except Exception as e:
        raise e
